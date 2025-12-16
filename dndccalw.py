import os
import sys
import pandas as pd
import numpy as np
from skopt import gp_minimize
from skopt.space import Real
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import PatternFill
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.linear_model import LinearRegression
import subprocess
import shutil
import copy
import logging
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import ctypes
import threading
import portalocker

# --------------------- Path Handling for PyInstaller ---------------------
def resource_path(relative_path):
    """Get the absolute path to a resource, works for dev and PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Global variables
TEMPLATE_BUTTON_WIDTH = 15
stop_calibration_flag = False
zoom_scale = 1.0  # Default zoom scale
calibration_thread = None

# --------------------- Light UI Theme Colors ---------------------
THEME = {
    "bg": "#f8f9fa",           # Light gray background
    "fg": "#212529",           # Dark text
    "primary": "#007bff",      # Blue - primary buttons
    "secondary": "#6c757d",    # Gray - secondary buttons
    "success": "#28a745",      # Green - start/positive actions
    "danger": "#dc3545",       # Red - stop/exit actions
    "warning": "#ffc107",      # Yellow - templates
    "info": "#17a2b8",         # Teal - info buttons
    "light": "#e9ecef",        # Light gray
    "dark": "#343a40",         # Dark gray
    "entry_bg": "#ffffff",     # White entry fields
    "entry_fg": "#212529",
    "log_bg": "#ffffff",       # White log panel
    "log_fg": "#212529",
    "card_bg": "#ffffff",      # White card background
    "header_bg": "#17a2b8",    # Blue header background
    "progress_fg": "#ffffff",  # White progress text
    "border": "#dee2e6"        # Border color
}

# --------------------- Stylish Button Function ---------------------
def create_styled_button(parent, text, command, style_type="primary", width=None, height=1):
    """Create a styled button with modern look"""

    styles = {
        "primary": {
            "bg": THEME["primary"],
            "fg": "white",
            "hover_bg": "#0056b3",
            "active_bg": "#004085"
        },
        "success": {
            "bg": THEME["success"],
            "fg": "white",
            "hover_bg": "#218838",
            "active_bg": "#1e7e34"
        },
        "danger": {
            "bg": THEME["danger"],
            "fg": "white",
            "hover_bg": "#c82333",
            "active_bg": "#bd2130"
        },
        "warning": {
            "bg": THEME["warning"],
            "fg": "#212529",
            "hover_bg": "#e0a800",
            "active_bg": "#d39e00"
        },
        "info": {
            "bg": THEME["info"],
            "fg": "white",
            "hover_bg": "#138496",
            "active_bg": "#117a8b"
        },
        "secondary": {
            "bg": THEME["secondary"],
            "fg": "white",
            "hover_bg": "#545b62",
            "active_bg": "#4e555b"
        }
    }

    style = styles.get(style_type, styles["primary"])

    btn = tk.Button(
        parent,
        text=text,
        command=command,
        bg=style["bg"],
        fg=style["fg"],
        font=("Segoe UI", 10, "bold"),
        relief="flat",
        bd=0,
        padx=20,
        pady=10,
        cursor="hand2",
        width=width,
        height=height
    )

    # Hover effects
    def on_enter(e):
        if btn['state'] == 'normal':
            btn.config(bg=style["hover_bg"])

    def on_leave(e):
        if btn['state'] == 'normal':
            btn.config(bg=style["bg"])

    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)

    return btn

# --------------------- Zoom Functions ---------------------
def zoom_in(event=None):
    global zoom_scale
    zoom_scale = min(2.0, zoom_scale + 0.1)  # Max zoom 2.0x
    apply_zoom()
    log_message(f"Zoom increased to {zoom_scale:.1f}x")

def zoom_out(event=None):
    global zoom_scale
    zoom_scale = max(0.5, zoom_scale - 0.1)  # Min zoom 0.5x
    apply_zoom()
    log_message(f"Zoom decreased to {zoom_scale:.1f}x")

def reset_zoom(event=None):
    global zoom_scale
    zoom_scale = 1.0
    apply_zoom()
    log_message("Zoom reset to 100%")

def apply_zoom():
    global zoom_scale
    try:
        # Update font sizes based on zoom scale
        default_font_size = 10
        scaled_font_size = max(8, int(default_font_size * zoom_scale))
        text_font_size = max(8, int(9 * zoom_scale))
        title_font_size = max(12, int(14 * zoom_scale))
        header_font_size = max(16, int(20 * zoom_scale))

        # Update fonts for all widgets
        update_all_widget_fonts(root, scaled_font_size, text_font_size, title_font_size, header_font_size)

        # Update progress bar length
        progress_bar.config(length=int(600 * zoom_scale))

    except Exception as e:
        log_message(f"Error applying zoom: {e}")

def update_all_widget_fonts(widget, scaled_font_size, text_font_size, title_font_size, header_font_size):
    """Recursively update fonts for all widgets"""
    try:
        if isinstance(widget, tk.Label):
            current_text = widget.cget("text")
            if "🌱" in current_text:  # Header
                widget.config(font=("Segoe UI", header_font_size, "bold"))
            elif "📊" in current_text or "📝" in current_text:  # Card titles
                widget.config(font=("Segoe UI", title_font_size, "bold"))
            else:
                widget.config(font=("Segoe UI", scaled_font_size))

        elif isinstance(widget, tk.Button):
            widget.config(font=("Segoe UI", scaled_font_size, "bold"))

        elif isinstance(widget, tk.Entry):
            widget.config(font=("Segoe UI", scaled_font_size))

        elif isinstance(widget, ttk.Combobox):
            widget.config(font=("Segoe UI", scaled_font_size))

        elif isinstance(widget, tk.Text):
            widget.config(font=("Consolas", text_font_size))

        # Recursively update child widgets
        if isinstance(widget, (tk.Frame, ttk.Frame, tk.Toplevel, tk.Tk)):
            for child in widget.winfo_children():
                update_all_widget_fonts(child, scaled_font_size, text_font_size, title_font_size, header_font_size)

    except Exception as e:
        pass  # Skip widgets that don't support font changes

# --------------------- Keyboard Shortcuts ---------------------
def setup_keyboard_shortcuts():
    """Setup keyboard shortcuts for zoom and other functions"""
    root.bind("<Control-plus>", zoom_in)      # Ctrl + +
    root.bind("<Control-equal>", zoom_in)     # Ctrl + = (same as +)
    root.bind("<Control-minus>", zoom_out)    # Ctrl + -
    root.bind("<Control-0>", reset_zoom)      # Ctrl + 0
    root.bind("<F1>", lambda e: start_calibration())  # F1 to start calibration
    root.bind("<Escape>", lambda e: exit_application())  # Esc to exit

# --------------------- Configuration ---------------------
ROOT_FOLDER = r"C:\DNDC"
DNDC_EXECUTABLE = os.path.join(ROOT_FOLDER, "DNDC95.exe")
OUTPUT_DIR = os.path.join(ROOT_FOLDER, "output_files")
MODELED_YIELD_CSV = os.path.join(OUTPUT_DIR, "Record", "Batch", "Case1-stc", "Multi_year_summary.csv")
MODELED_SOIL_CLIMATE_CSV = os.path.join(OUTPUT_DIR, "Record", "Batch", "Case1-stc", "Day_SoilClimate_1.csv")
MODELED_CLIMATE_CSV = os.path.join(OUTPUT_DIR, "Record", "Batch", "Case1-stc", "Day_Climate_1.csv")
MODELED_NEE_CSV = os.path.join(OUTPUT_DIR, "Record", "Batch", "Case1-stc", "Day_SoilC_1.csv")
HARDCODED_OBSERVED_YIELD_CSV = os.path.join(ROOT_FOLDER, "observed_yield.csv")
HARDCODED_PARAM_CSV = os.path.join(ROOT_FOLDER, "parm.csv")

# Soil temperature and moisture depth configurations
SOIL_TEMP_DEPTHS = {
    "1cm": 5, "5cm": 6, "10cm": 7, "20cm": 8, "30cm": 9, "40cm": 10,
    "50cm": 11, "60cm": 12, "70cm": 13, "80cm": 14, "90cm": 15, "100cm": 16,
    "110cm": 17, "120cm": 18, "130cm": 19, "140cm": 20, "150cm": 21,
    "160cm": 22, "170cm": 23, "180cm": 24, "190cm": 25, "200cm": 26
}

SOIL_MOISTURE_DEPTHS = {
    "1cm": 27, "5cm": 28, "10cm": 29, "20cm": 30, "30cm": 31, "40cm": 32,
    "50cm": 33, "60cm": 34, "70cm": 35, "80cm": 36, "90cm": 37, "100cm": 38,
    "110cm": 39, "120cm": 40, "130cm": 41, "140cm": 42, "150cm": 43,
    "160cm": 44, "170cm": 45, "180cm": 46, "190cm": 47, "200cm": 48
}

# Logging setup
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --------------------- Utility Functions ---------------------
error_dialog_shown = False

def show_error(message):
    global error_dialog_shown
    if not error_dialog_shown:
        error_dialog_shown = True
        messagebox.showerror("Error", message)
        error_dialog_shown = False

def log_message(message):
    log_display.insert(tk.END, message + "\n")
    log_display.see(tk.END)
    root.update()

def check_file_exists(file_path):
    if not os.path.exists(file_path):
        log_message(f"Error: The file '{file_path}' does not exist.")
        return False
    if not os.access(file_path, os.R_OK):
        log_message(f"Error: The file '{file_path}' is not readable.")
        return False
    return True

def read_dnd_file(dnd_path):
    try:
        with open(dnd_path, 'r') as file:
            return file.readlines()
    except Exception as e:
        log_message(f"Failed to read .dnd file: {e}")
        return []

def write_dnd_file(dnd_path, lines):
    try:
        with open(dnd_path, 'w') as file:
            file.writelines(lines)
    except Exception as e:
        log_message(f"Failed to write .dnd file: {e}")

def read_param_ranges(param_csv):
    try:
        df = pd.read_csv(param_csv)
        if not all(col in df.columns for col in ["parameter_name", "min", "max", "line_number"]):
            raise ValueError("Missing required columns in CSV")
        return df
    except Exception as e:
        log_message(f"Failed to read parameters: {e}")
        return pd.DataFrame()

def update_parameters(lines, param_values, param_ranges_df):
    updated_lines = copy.deepcopy(lines)
    for param, value in param_values.items():
        param_row = param_ranges_df[param_ranges_df["parameter_name"] == param]
        if not param_row.empty:
            line_idx = param_row["line_number"].values[0] - 1
            if line_idx < len(updated_lines):
                parts = updated_lines[line_idx].strip().split()
                if len(parts) >= 2:
                    parts[1] = f"{value:.6f}"
                    updated_lines[line_idx] = ' '.join(parts) + '\n'
    return updated_lines

def run_dndc(output_dir):
    try:
        result = subprocess.run([
            DNDC_EXECUTABLE,
            "-root", ROOT_FOLDER,
            "-output", output_dir,
            "-s", BATCH_FILE,
            "-daily", "1"
        ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        log_message(f"DNDC output: {result.stdout}")
    except subprocess.CalledProcessError as e:
        log_message(f"DNDC run failed: {e.stderr}")
        raise

# --------------------- Data Reading Functions ---------------------
def read_yield_data(modeled_path, observed_path):
    try:
        modeled_df = pd.read_csv(modeled_path, skiprows=5, usecols=[0, 2], names=['Year', 'Yield'], header=None)
        modeled_df.columns = ['Year', 'Yield_MOD']
        observed_df = pd.read_csv(observed_path, skiprows=2, usecols=[0, 1], names=['Year', 'Yield'], header=None)
        observed_df.columns = ['Year', 'Yield_OBS']
        return modeled_df, observed_df
    except Exception as e:
        log_message(f"Error reading yield data: {e}")
        return pd.DataFrame(), pd.DataFrame()

def read_soil_temp_data(modeled_path, observed_path, depth):
    try:
        if isinstance(depth, (int, float)):
            depth = next((k for k, v in SOIL_TEMP_DEPTHS.items() if v == int(depth)), None)
            if depth is None:
                raise ValueError(f"Invalid numeric depth: {depth}")

        if depth not in SOIL_TEMP_DEPTHS:
            raise ValueError(f"Invalid depth '{depth}'")

        col_idx = SOIL_TEMP_DEPTHS[depth]
        modeled_df = pd.read_csv(modeled_path, skiprows=4, header=0)
        if col_idx >= len(modeled_df.columns):
            raise IndexError(f"Column index {col_idx} out of range")

        modeled_df = modeled_df.iloc[:, [0, 1, col_idx]]
        modeled_df.columns = ['Year', 'Day', 'SoilTemp_MOD']

        observed_df = pd.read_csv(observed_path, skiprows=2, header=None,
                                  usecols=[0, 1, 2], names=['Year', 'Day', 'SoilTemp_OBS'])

        modeled_df = modeled_df.dropna()
        observed_df = observed_df.dropna()
        if modeled_df.empty or observed_df.empty:
            log_message("Warning: Empty DataFrame after cleaning")
            return pd.DataFrame(), pd.DataFrame()

        return modeled_df, observed_df
    except Exception as e:
        log_message(f"Error reading {depth} data: {str(e)}")
        return pd.DataFrame(), pd.DataFrame()

def read_soil_moisture_data(modeled_path, observed_path, depth):
    try:
        if isinstance(depth, (int, float)):
            depth = next(
                (k for k, v in SOIL_MOISTURE_DEPTHS.items() if v == int(depth) or int(depth) == int(k.replace("cm", ""))),
                None
            )
            if depth is None:
                log_message(f"Invalid depth '{depth}'")
                return pd.DataFrame(), pd.DataFrame()

        if depth not in SOIL_MOISTURE_DEPTHS:
            log_message(f"Invalid depth '{depth}'")
            return pd.DataFrame(), pd.DataFrame()

        col_idx = SOIL_MOISTURE_DEPTHS[depth]
        modeled_df = pd.read_csv(modeled_path, skiprows=4, header=0)
        if col_idx >= len(modeled_df.columns):
            raise IndexError(f"Column index {col_idx} out of range")

        modeled_df = modeled_df.iloc[:, [0, 1, col_idx]]
        modeled_df.columns = ['Year', 'Day', 'SoilMoisture_MOD']

        observed_df = pd.read_csv(observed_path, skiprows=2, header=None,
                                  usecols=[0, 1, 2], names=['Year', 'Day', 'SoilMoisture_OBS'])

        for df in [modeled_df, observed_df]:
            df['Year'] = pd.to_numeric(df['Year'], errors='coerce').fillna(0).astype(int)
            df['Day'] = pd.to_numeric(df['Day'], errors='coerce').fillna(0).astype(int)
            df.iloc[:, 2] = pd.to_numeric(df.iloc[:, 2], errors='coerce')

        modeled_df = modeled_df.dropna()
        observed_df = observed_df.dropna()

        if modeled_df.empty or observed_df.empty:
            log_message("No valid data after cleaning")
            return pd.DataFrame(), pd.DataFrame()

        return modeled_df, observed_df
    except Exception as e:
        log_message(f"Error reading soil moisture data for depth {depth}: {str(e)}")
        return pd.DataFrame(), pd.DataFrame()

def read_et_data(modeled_path, observed_path):
    try:
        modeled_df = pd.read_csv(modeled_path, skiprows=[0], header=1)
        modeled_df = modeled_df.iloc[:, [0, 1, 9]]
        modeled_df.columns = ['Year', 'Day', 'ET_MOD']

        modeled_df['Year'] = modeled_df['Year'].astype(int)
        modeled_df['Day'] = modeled_df['Day'].astype(int)
        modeled_df['ET_MOD'] = pd.to_numeric(modeled_df['ET_MOD'], errors='coerce')

        observed_df = pd.read_csv(observed_path, skiprows=2, header=None,
                                usecols=[0, 1, 2], names=['Year', 'Day', 'ET_OBS'])

        observed_df['Year'] = observed_df['Year'].astype(int)
        observed_df['Day'] = observed_df['Day'].astype(int)
        observed_df['ET_OBS'] = pd.to_numeric(observed_df['ET_OBS'], errors='coerce')

        modeled_df = modeled_df.dropna()
        observed_df = observed_df.dropna()

        if modeled_df.empty or observed_df.empty:
            log_message("No valid data after cleaning")
            return pd.DataFrame(), pd.DataFrame()

        return modeled_df, observed_df
    except Exception as e:
        log_message(f"Error reading ET data: {str(e)}")
        return pd.DataFrame(), pd.DataFrame()

def read_nee_data(modeled_path, observed_path):
    try:
        modeled_df = pd.read_csv(modeled_path, skiprows=1)
        modeled_df = modeled_df.iloc[:, [0, 1, 42]]
        modeled_df.columns = ['Year', 'Day', 'NEE_MOD']

        modeled_df['Year'] = modeled_df['Year'].astype(int)
        modeled_df['Day'] = modeled_df['Day'].astype(int)
        modeled_df['NEE_MOD'] = pd.to_numeric(modeled_df['NEE_MOD'], errors='coerce')

        observed_df = pd.read_csv(observed_path, skiprows=2, usecols=[0, 1, 2],
                                names=['Year', 'Day', 'NEE'], header=None)
        observed_df.columns = ['Year', 'Day', 'NEE_OBS']

        observed_df['Year'] = observed_df['Year'].astype(int)
        observed_df['Day'] = observed_df['Day'].astype(int)
        observed_df['NEE_OBS'] = pd.to_numeric(observed_df['NEE_OBS'], errors='coerce')

        return modeled_df, observed_df
    except Exception as e:
        log_message(f"Error reading NEE data: {e}")
        return pd.DataFrame(), pd.DataFrame()

# --------------------- Metrics Calculation ---------------------
def calculate_metrics(y_true, y_pred):
    r2 = r2_score(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    mae = mean_absolute_error(y_true, y_pred)
    mbe = np.mean(y_pred - y_true)
    try:
        X = np.array(y_pred).reshape(-1, 1)
        model = LinearRegression()
        model.fit(X, y_true)
        lr_r2 = model.score(X, y_true)
    except Exception as e:
        log_message(f"Error calculating linear regression R²: {e}")
        lr_r2 = np.nan
    return {
        'R2': r2,
        'LR_R2': lr_r2,
        'RMSE': rmse,
        'MAE': mae,
        'MBE': mbe
    }

def match_and_evaluate(modeled_df, observed_df, target_var):
    if modeled_df.empty or observed_df.empty:
        log_message(f"Error: Empty DataFrame provided for {target_var}.")
        return None, pd.DataFrame()

    for col in ['Year', 'Day']:
        if col in modeled_df.columns:
            modeled_df[col] = modeled_df[col].astype(int)
        if col in observed_df.columns:
            observed_df[col] = observed_df[col].astype(int)

    if target_var == "Yield":
        if 'Yield_MOD' in modeled_df.columns:
            modeled_df['Yield_MOD'] = modeled_df['Yield_MOD'] / 0.4
        merge_on = ['Year']
    else:
        merge_on = ['Year', 'Day']

    merged_df = pd.merge(modeled_df, observed_df, on=merge_on, how='inner')

    if merged_df.empty:
        log_message(f"No matching data after merging modeled and observed {target_var} data.")
        return None, pd.DataFrame()

    observed_col = f"{target_var}_OBS"
    merged_df.dropna(subset=[observed_col], inplace=True)

    if merged_df.empty:
        log_message(f"No valid observed {target_var} values after dropping NaN.")
        return None, pd.DataFrame()

    y_true = merged_df[observed_col]
    y_pred = merged_df[f"{target_var}_MOD"]
    metrics = calculate_metrics(y_true, y_pred)

    return metrics, merged_df

# --------------------- Optimization Core ---------------------
def objective_function(params, param_ranges_df, lines, target_var, depth=None):
    log_message(f"Objective function called for {target_var} with depth: {depth}")

    param_dict = dict(zip(param_ranges_df['parameter_name'], params))
    updated_lines = update_parameters(lines, param_dict, param_ranges_df)
    write_dnd_file(DND_FILE, updated_lines)
    run_dndc(OUTPUT_DIR)

    if target_var == "Yield":
        modeled_df, observed_df = read_yield_data(MODELED_YIELD_CSV, OBSERVED_YIELD_CSV)
    elif target_var == "SoilTemp":
        depth_col = SOIL_TEMP_DEPTHS[depth]
        modeled_df, observed_df = read_soil_temp_data(MODELED_SOIL_CLIMATE_CSV, OBSERVED_YIELD_CSV, depth_col)
    elif target_var == "SoilMoisture":
        depth_col = SOIL_MOISTURE_DEPTHS[depth]
        modeled_df, observed_df = read_soil_moisture_data(MODELED_SOIL_CLIMATE_CSV, OBSERVED_YIELD_CSV, depth_col)
    elif target_var == "ET":
        modeled_df, observed_df = read_et_data(MODELED_CLIMATE_CSV, OBSERVED_YIELD_CSV)
    elif target_var == "NEE":
        modeled_df, observed_df = read_nee_data(MODELED_NEE_CSV, OBSERVED_YIELD_CSV)
    else:
        log_message(f"Unknown target variable: {target_var}")
        return np.inf

    yield_metrics, _ = match_and_evaluate(modeled_df, observed_df, target_var)
    if yield_metrics is None:
        return np.inf

    log_message(f"Calculated RMSE: {yield_metrics['RMSE']:.4f} for depth {depth}")
    return yield_metrics['RMSE']

def bayesian_optimization(param_ranges, param_ranges_df, lines, target_var, depth=None):
    log_message(f"\nStarting Bayesian Optimization for {target_var}{f' at {depth}' if depth else ''}...")

    all_results = []
    best_rmse = np.inf
    best_params = None
    best_merged = None
    best_metrics = None
    best_iteration = 0
    iteration_counter = 0
    total_iterations = int(iterations_entry.get())

    def callback(res):
        nonlocal best_rmse, best_params, best_merged, best_metrics, best_iteration, iteration_counter

        iteration_counter += 1

        global stop_calibration_flag
        if stop_calibration_flag:
            log_message("Calibration stopped by user request.")
            raise StopIteration("Calibration stopped by user.")

        try:
            current_rmse = res.func_vals[-1]
            params = res.x_iters[-1]

            depth_col = None
            if target_var == "SoilTemp":
                depth_col = SOIL_TEMP_DEPTHS.get(depth, None)
            elif target_var == "SoilMoisture":
                depth_col = SOIL_MOISTURE_DEPTHS.get(depth, None)

            if target_var in ["SoilTemp", "SoilMoisture"] and depth_col is None:
                log_message(f"Error: depth_col is None for {target_var}. Skipping iteration.")
                return

            if target_var == "Yield":
                modeled_df, observed_df = read_yield_data(MODELED_YIELD_CSV, OBSERVED_YIELD_CSV)
            elif target_var == "SoilTemp":
                modeled_df, observed_df = read_soil_temp_data(MODELED_SOIL_CLIMATE_CSV, OBSERVED_YIELD_CSV, depth_col)
            elif target_var == "SoilMoisture":
                modeled_df, observed_df = read_soil_moisture_data(MODELED_SOIL_CLIMATE_CSV, OBSERVED_YIELD_CSV, depth_col)
            elif target_var == "ET":
                modeled_df, observed_df = read_et_data(MODELED_CLIMATE_CSV, OBSERVED_YIELD_CSV)
            elif target_var == "NEE":
                modeled_df, observed_df = read_nee_data(MODELED_NEE_CSV, OBSERVED_YIELD_CSV)
            else:
                log_message(f"Unknown target variable: {target_var}")
                return

            if modeled_df.empty or observed_df.empty:
                log_message(f"Error: Modeled or observed data is empty for {target_var}. Skipping iteration.")
                return

            yield_metrics, merged_df = match_and_evaluate(modeled_df, observed_df, target_var)

            if yield_metrics is None:
                log_message(f"Error: Metrics could not be calculated for {target_var}. Skipping iteration.")
                return

            iteration_result = {
                "Iteration": iteration_counter,
                "Parameters": params,
                "Metrics": yield_metrics,
                "Merged_Data": merged_df
            }
            all_results.append(iteration_result)

            if current_rmse < best_rmse:
                best_rmse = current_rmse
                best_params = params
                best_merged = merged_df
                best_metrics = yield_metrics
                best_iteration = iteration_counter
                log_message(f"Iteration {iteration_counter}: New best RMSE {current_rmse:.2f} for depth {depth}")

            log_message(f"Iteration {iteration_counter} Metrics: R²={yield_metrics['R2']:.4f}, RMSE={yield_metrics['RMSE']:.4f}")
            log_message("\nIteration {}".format(iteration_counter))
            log_message("Parameters:")
            for param, value in zip(param_ranges_df['parameter_name'], params):
                log_message(f"  {param}: {value:.6f}")
            log_message(f"{target_var} Metrics:")
            log_message(f"  R²: {yield_metrics['R2']:.4f}, RMSE: {yield_metrics['RMSE']:.4f}, MAE: {yield_metrics['MAE']:.4f}")

            progress_percentage = (iteration_counter / total_iterations) * 100
            root.after(0, lambda: progress_bar.configure(value=progress_percentage))
            root.after(0, lambda: progress_label.config(text=f"{progress_percentage:.1f}%"))
            root.update_idletasks()

            backup_dir = os.path.join(os.path.dirname(DND_FILE), "dnd_backups")
            os.makedirs(backup_dir, exist_ok=True)

            current_params = res.x_iters[-1]
            param_names = param_ranges_df['parameter_name'].tolist()
            param_values = [f"{name}_{value:.4f}" for name, value in zip(param_names, current_params)]
            backup_filename = f"iter_{iteration_counter}.dnd"
            backup_path = os.path.join(backup_dir, backup_filename)

            try:
                shutil.copy(DND_FILE, backup_path)
                log_message(f"Backup created: {backup_filename}")
            except Exception as e:
                log_message(f"Backup failed: {e}")

        except Exception as e:
            log_message(f"Unexpected error during iteration {iteration_counter}: {e}")

    try:
        result = gp_minimize(
            lambda p: objective_function(p, param_ranges_df, lines, target_var, depth),
            param_ranges,
            n_calls=total_iterations,
            callback=callback,
            random_state=42,
            n_jobs=1
        )
        log_message("Bayesian Optimization finished")
    except StopIteration:
        log_message("Calibration stopped by user. Compiling results...")
    finally:
        save_results(all_results, best_params, best_metrics, best_merged, param_ranges_df, target_var, depth)
        log_message("Results compiled and saved.")
        return all_results, best_params, best_merged, best_metrics

# --------------------- Results Handling ---------------------
def save_results(all_results, best_params, best_metrics, best_merged, param_ranges_df, target_var, depth=None):
    if target_var == "Yield":
        output_file = os.path.join(ROOT_FOLDER, "yield_calibration_results.xlsx")
    elif target_var == "SoilTemp":
        output_file = os.path.join(ROOT_FOLDER, f"soil_temp_{depth}_results.xlsx")
    elif target_var == "SoilMoisture":
        output_file = os.path.join(ROOT_FOLDER, f"soil_moisture_{depth}_results.xlsx")
    elif target_var == "ET":
        output_file = os.path.join(ROOT_FOLDER, "et_calibration_results.xlsx")
    elif target_var == "NEE":
        output_file = os.path.join(ROOT_FOLDER, "nee_calibration_results.xlsx")

    wb = Workbook()

    sheet_all_iterations = wb.active
    sheet_all_iterations.title = "All Iterations"

    headers = ["Iteration"] + param_ranges_df['parameter_name'].tolist() + ["R2", "LR_R2", "RMSE", "MAE", "MBE"]
    sheet_all_iterations.append(headers)

    best_iter_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    best_iteration_num = all_results[-1]["Iteration"] if all_results else 0

    for result in all_results:
        row_data = [result["Iteration"]] + list(result["Parameters"]) + [
            result["Metrics"]["R2"],
            result["Metrics"]["LR_R2"],
            result["Metrics"]["RMSE"],
            result["Metrics"]["MAE"],
            result["Metrics"]["MBE"]
        ]
        sheet_all_iterations.append(row_data)

        if result["Iteration"] == best_iteration_num:
            for cell in sheet_all_iterations[sheet_all_iterations.max_row]:
                cell.fill = best_iter_fill

    sheet_best_iteration = wb.create_sheet("Best Iteration")
    sheet_best_iteration.append([
        "Iteration",
        *param_ranges_df['parameter_name'].tolist(),
        "R2", "LR_R2", "RMSE", "MAE", "MBE"
    ])
    if all_results:
        sheet_best_iteration.append([
            all_results[-1]["Iteration"],
            *best_params,
            best_metrics['R2'],
            best_metrics['LR_R2'],
            best_metrics['RMSE'],
            best_metrics['MAE'],
            best_metrics['MBE']
        ])

    ws_data = wb.create_sheet("Data Comparison")
    if 'Day' in best_merged.columns:
        ws_data.append(["Year", "Day", "Observed", "Modeled"])
        for _, row in best_merged.iterrows():
            ws_data.append([row['Year'], row['Day'],
                          row[f"{target_var}_OBS"],
                          row[f"{target_var}_MOD"]])
    else:
        ws_data.append(["Year", "Observed", "Modeled"])
        for _, row in best_merged.iterrows():
            ws_data.append([row['Year'],
                          row[f"{target_var}_OBS"],
                          row[f"{target_var}_MOD"]])

    if 'Day' in best_merged.columns:
        chart = LineChart()
        chart.title = f"{target_var}{f' at {depth}' if depth else ''} Calibration Results"
        chart.style = 12
        data = Reference(ws_data, min_col=3, max_col=4, min_row=1, max_row=len(best_merged)+1)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=len(best_merged)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.y_axis.title = f"{target_var} Value"
        chart.x_axis.title = "Date"
    else:
        chart = BarChart()
        chart.type = "col"
        chart.style = 12
        chart.title = f"{target_var}{f' at {depth}' if depth else ''} Calibration Results"
        chart.y_axis.title = f"{target_var} Value"
        chart.x_axis.title = "Year"
        data = Reference(ws_data, min_col=2, max_col=3, min_row=1, max_row=len(best_merged)+1)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=len(best_merged)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

    if len(chart.series) >= 2:
        obs_color = "4DC4E3"
        mod_color = "FF6F61"
        chart.series[0].graphicalProperties.solidFill = obs_color
        chart.series[1].graphicalProperties.solidFill = mod_color
        chart.series[0].tx.v = "Observed"
        chart.series[1].tx.v = "Modeled"
        if isinstance(chart, BarChart):
            chart.gapWidth = 50

    ws_data.add_chart(chart, "E5")

    ws_all_modeled = wb.create_sheet("All Iterations Data")
    headers = ["Year", "Day"] + [f"Iteration {i+1}" for i in range(len(all_results))]
    ws_all_modeled.append(headers)

    data_dict = {}
    for iteration_idx, result in enumerate(all_results):
        merged_df = result["Merged_Data"]
        for _, row in merged_df.iterrows():
            year = row['Year']
            day = row.get('Day', '')
            key = (year, day)

            if key not in data_dict:
                data_dict[key] = [year, day] + [None] * len(all_results)

            data_dict[key][2 + iteration_idx] = row[f"{target_var}_MOD"]

    for row_data in data_dict.values():
        ws_all_modeled.append(row_data)

    wb.save(output_file)
    log_message(f"Results saved to '{output_file}'.")

# --------------------- Main Workflow ---------------------
def start_calibration():
    global BATCH_FILE, DND_FILE, OBSERVED_YIELD_CSV, PARAM_CSV, calibration_thread, stop_calibration_flag
    stop_calibration_flag = False

    if calibration_thread and calibration_thread.is_alive():
        calibration_thread.join(timeout=5)
        if calibration_thread.is_alive():
            log_message("Previous calibration thread did not terminate gracefully.")
            return

    BATCH_FILE = batch_file_entry.get()
    DND_FILE = dnd_file_entry.get()
    OBSERVED_YIELD_CSV = observed_csv_entry.get()
    PARAM_CSV = param_csv_entry.get()

    if not all([BATCH_FILE, DND_FILE, OBSERVED_YIELD_CSV, PARAM_CSV]):
        log_message("Error: Please fill all fields.")
        return

    target_var = target_var_combo.get()
    depth = None
    if target_var in ["SoilTemp", "SoilMoisture"]:
        depth = depth_combo.get()
        if not depth:
            log_message(f"Error: Please select a depth for {target_var} calibration.")
            return

    log_message(f"Starting {target_var}{f' at {depth}' if depth else ''} calibration process...")

    try:
        progress_bar.configure(value=0)
        progress_label.config(text="0.0%")
        calibration_thread = threading.Thread(
            target=calibrate_variable,
            args=(target_var, depth),
            daemon=True
        )
        calibration_thread.start()
    except Exception as e:
        log_message(f"Error during calibration: {e}")
        show_error(f"Error during calibration: {e}")

def calibrate_variable(target_var, depth=None):
    global stop_calibration_flag
    stop_calibration_flag = False

    log_message(f"Calibration started for {target_var}{f' at {depth}' if depth else ''}.")
    backup_path = DND_FILE + ".backup"
    shutil.copy(DND_FILE, backup_path)
    log_message(f"Original .dnd file backed up to '{backup_path}'.")

    try:
        lines = read_dnd_file(DND_FILE)
        param_ranges_df = read_param_ranges(PARAM_CSV)
        param_ranges = [(row['min'], row['max']) for _, row in param_ranges_df.iterrows()]

        all_results = []
        best_params = None
        best_metrics = None
        best_merged = pd.DataFrame()

        results = bayesian_optimization(
            param_ranges, param_ranges_df, lines, target_var, depth
        )

        all_results, best_params, best_merged, best_metrics = results if results else ([], [None]*len(param_ranges_df), pd.DataFrame(), {})

        if best_params and not any(p is None for p in best_params):
            save_results(all_results, best_params, best_metrics, best_merged,
                        param_ranges_df, target_var, depth)
            log_message(f"Calibration complete. Results saved.")
        else:
            log_message("Calibration completed, but no best parameters were found.")
    except Exception as e:
        log_message(f"An error occurred during calibration: {e}")
        show_error(f"An error occurred during calibration: {e}")
    finally:
        shutil.copy(backup_path, DND_FILE)
        log_message(f"Original .dnd file restored from '{backup_path}'.")

def stop_calibration():
    global stop_calibration_flag
    if calibration_thread and calibration_thread.is_alive():
        response = messagebox.askyesno(
            "Stop Calibration",
            "Are you sure you want to stop the calibration process? Results will be saved up to the current iteration."
        )
        if response:
            log_message("Stopping calibration after the current iteration...")
            stop_calibration_flag = True
    else:
        log_message("No active calibration process to stop.")

# --------------------- UI Functions ---------------------
def on_target_var_change(event):
    target_var = target_var_combo.get()
    if target_var in ["SoilTemp", "SoilMoisture"]:
        depth_label.grid()
        depth_combo.grid()
        if target_var == "SoilTemp":
            depth_combo['values'] = list(SOIL_TEMP_DEPTHS.keys())
        else:
            depth_combo['values'] = list(SOIL_MOISTURE_DEPTHS.keys())
        depth_combo.current(0)
        log_message(f"Updated depth options for {target_var}. Selected depth: {depth_combo.get()}")
    else:
        depth_label.grid_remove()
        depth_combo.grid_remove()

def browse_file(entry_widget, file_type, title):
    filename = filedialog.askopenfilename(filetypes=[(f"{file_type[1:].upper()} files", file_type)], title=title)
    if filename:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, filename)

def download_observed_template():
    try:
        save_dir = filedialog.askdirectory(title="Select Directory to Save Template")
        if not save_dir:
            return
        target_var = target_var_combo.get()
        if target_var == "Yield":
            template = """Year,Value
1,5000
2,5200
3,5100
"""
        else:
            template = """Year,Day,Value
1,1,5000
1,2,5200
1,3,5100
"""
        with open(os.path.join(save_dir, f"observed_{target_var.lower()}_template.csv"), "w") as f:
            f.write(template)
        log_message(f"Template saved to '{save_dir}' as 'observed_{target_var.lower()}_template.csv'.")
    except Exception as e:
        log_message(f"Error saving template: {e}")

def download_param_template():
    try:
        save_dir = filedialog.askdirectory(title="Select Directory to Save Template")
        if not save_dir:
            return
        template = """parameter_name,min,max,line_number
param1,0.1,1.0,10
param2,0.5,2.0,15
param3,0.0,0.9,20
"""
        with open(os.path.join(save_dir, "parameter_template.csv"), "w") as f:
            f.write(template)
        log_message(f"Parameter template saved to '{save_dir}' as 'parameter_template.csv'.")
    except Exception as e:
        log_message(f"Error saving parameter template: {e}")

def open_results_directory():
    try:
        target_var = target_var_combo.get()
        depth = depth_combo.get() if target_var in ["SoilTemp", "SoilMoisture"] else None

        if target_var == "Yield":
            results_file = os.path.join(ROOT_FOLDER, "yield_calibration_results.xlsx")
        elif target_var == "SoilTemp":
            results_file = os.path.join(ROOT_FOLDER, f"soil_temp_{depth}_results.xlsx")
        elif target_var == "SoilMoisture":
            results_file = os.path.join(ROOT_FOLDER, f"soil_moisture_{depth}_results.xlsx")
        elif target_var == "ET":
            results_file = os.path.join(ROOT_FOLDER, "et_calibration_results.xlsx")
        elif target_var == "NEE":
            results_file = os.path.join(ROOT_FOLDER, "nee_calibration_results.xlsx")
        else:
            log_message("Error: Unknown target variable selected")
            return

        if os.path.exists(results_file):
            os.startfile(results_file)
            log_message(f"Opened results file: {results_file}")
        else:
            log_message(f"Results file not found: {results_file}")
            os.startfile(ROOT_FOLDER)
    except Exception as e:
        log_message(f"Error opening results: {e}")
        os.startfile(ROOT_FOLDER)

def exit_application():
    global calibration_thread
    if calibration_thread and calibration_thread.is_alive():
        response = messagebox.askyesno(
            "Exit Application",
            "A calibration process is currently running. Are you sure you want to exit?"
        )
        if not response:
            return
        calibration_thread.join(timeout=5)
        if calibration_thread.is_alive():
            log_message("Calibration thread did not terminate gracefully. Forcing exit.")
    root.destroy()

def is_already_running():
    lock_file_path = os.path.join(ROOT_FOLDER, "dnccalibration.lock")
    try:
        lock_file = open(lock_file_path, 'w')
        portalocker.lock(lock_file, portalocker.LOCK_EX | portalocker.LOCK_NB)
        return False
    except portalocker.exceptions.LockException:
        return True
    finally:
        if 'lock_file' in locals():
            lock_file.close()
            os.remove(lock_file_path)

# --------------------- Main UI Setup ---------------------
def create_ui():
    global root, log_display, batch_file_entry, dnd_file_entry, observed_csv_entry
    global param_csv_entry, iterations_entry, progress_bar, progress_label
    global target_var_combo, depth_combo, depth_label

    root = tk.Tk()
    root.title("DNDC Multi-Variable Calibration Tool")
    root.geometry("1200x800")
    root.resizable(True, True)
    root.config(bg=THEME["bg"])

    # Declare DPI awareness for Windows
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass

    # Setup keyboard shortcuts
    setup_keyboard_shortcuts()

    # Header with modern design
    header_frame = tk.Frame(root, bg=THEME["header_bg"], height=80)
    header_frame.pack(fill=tk.X, pady=(0, 20))
    header_frame.pack_propagate(False)

    header_label = tk.Label(header_frame,
                          text="🌱 DNDC Multi-Variable Calibration Tool",
                          font=("Segoe UI", 20, "bold"),
                          bg=THEME["header_bg"],
                          fg="white",
                          pady=30)
    header_label.pack(expand=True)

    # Main container
    main_container = tk.Frame(root, bg=THEME["bg"], padx=30, pady=20)
    main_container.pack(fill=tk.BOTH, expand=True)

    # Settings Card
    settings_card = tk.Frame(main_container, bg=THEME["card_bg"], relief="solid", bd=1, padx=20, pady=20)
    settings_card.pack(fill=tk.X, pady=(0, 20))

    # Card title
    card_title = tk.Label(settings_card, text="📊 Calibration Settings",
                         font=("Segoe UI", 14, "bold"),
                         bg=THEME["card_bg"], fg=THEME["fg"])
    card_title.grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 15))

    # Target Variable Selection
    tk.Label(settings_card, text="Target Variable:", bg=THEME["card_bg"], fg=THEME["fg"],
            font=("Segoe UI", 10)).grid(row=1, column=0, padx=5, pady=8, sticky="w")
    target_var_combo = ttk.Combobox(settings_card, values=["Yield", "SoilTemp", "SoilMoisture", "ET", "NEE"],
                                   font=("Segoe UI", 10), width=15, state="readonly")
    target_var_combo.grid(row=1, column=1, padx=5, pady=8, sticky="w")
    target_var_combo.current(0)
    target_var_combo.bind("<<ComboboxSelected>>", on_target_var_change)

    # Depth Selection
    depth_label = tk.Label(settings_card, text="Depth:", bg=THEME["card_bg"], fg=THEME["fg"],
                          font=("Segoe UI", 10))
    depth_combo = ttk.Combobox(settings_card, font=("Segoe UI", 10), width=10, state="readonly")
    depth_label.grid(row=1, column=2, padx=5, pady=8, sticky="w")
    depth_combo.grid(row=1, column=3, padx=5, pady=8, sticky="w")
    depth_label.grid_remove()
    depth_combo.grid_remove()

    # File Input Section
    file_section = tk.Frame(settings_card, bg=THEME["card_bg"])
    file_section.grid(row=2, column=0, columnspan=6, sticky="w", pady=10)

    # Batch File
    tk.Label(file_section, text="Batch File:", bg=THEME["card_bg"], fg=THEME["fg"],
            font=("Segoe UI", 10)).grid(row=0, column=0, padx=5, pady=5, sticky="w")
    batch_file_entry = tk.Entry(file_section, width=50, bg=THEME["entry_bg"], fg=THEME["entry_fg"],
                               font=("Segoe UI", 10), relief="solid", bd=1)
    batch_file_entry.grid(row=0, column=1, columnspan=3, padx=5, pady=5, sticky="w")
    create_styled_button(file_section, "📁 Browse",
                       command=lambda: browse_file(batch_file_entry, "*.txt", "Select Batch File"),
                       style_type="secondary", width=10).grid(row=0, column=4, padx=5, pady=5)

    # DNDC File
    tk.Label(file_section, text=".dnd File:", bg=THEME["card_bg"], fg=THEME["fg"],
            font=("Segoe UI", 10)).grid(row=1, column=0, padx=5, pady=5, sticky="w")
    dnd_file_entry = tk.Entry(file_section, width=50, bg=THEME["entry_bg"], fg=THEME["entry_fg"],
                             font=("Segoe UI", 10), relief="solid", bd=1)
    dnd_file_entry.grid(row=1, column=1, columnspan=3, padx=5, pady=5, sticky="w")
    create_styled_button(file_section, "📁 Browse",
                       command=lambda: browse_file(dnd_file_entry, "*.dnd", "Select DNDC File"),
                       style_type="secondary", width=10).grid(row=1, column=4, padx=5, pady=5)

    # Observed CSV
    tk.Label(file_section, text="Observed CSV:", bg=THEME["card_bg"], fg=THEME["fg"],
            font=("Segoe UI", 10)).grid(row=2, column=0, padx=5, pady=5, sticky="w")
    observed_csv_entry = tk.Entry(file_section, width=50, bg=THEME["entry_bg"], fg=THEME["entry_fg"],
                                 font=("Segoe UI", 10), relief="solid", bd=1)
    observed_csv_entry.grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky="w")
    create_styled_button(file_section, "📁 Browse",
                       command=lambda: browse_file(observed_csv_entry, "*.csv", "Select Observed Data CSV"),
                       style_type="secondary", width=10).grid(row=2, column=4, padx=5, pady=5)

    # Obs. Template Button (on row 2)
    create_styled_button(file_section, "📥 Obs. Template", download_observed_template,
                       style_type="info", width=15).grid(row=2, column=5, padx=10, pady=5, sticky="w")

    # Parameter CSV
    tk.Label(file_section, text="Parameter CSV:", bg=THEME["card_bg"], fg=THEME["fg"],
            font=("Segoe UI", 10)).grid(row=3, column=0, padx=5, pady=5, sticky="w")
    param_csv_entry = tk.Entry(file_section, width=50, bg=THEME["entry_bg"], fg=THEME["entry_fg"],
                              font=("Segoe UI", 10), relief="solid", bd=1)
    param_csv_entry.grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky="w")
    create_styled_button(file_section, "📁 Browse",
                       command=lambda: browse_file(param_csv_entry, "*.csv", "Select Parameters CSV"),
                       style_type="secondary", width=10).grid(row=3, column=4, padx=5, pady=5)
    
    # Param. Template Button (on row 3)
    create_styled_button(file_section, "📥 Param. Template", download_param_template,
                       style_type="info", width=15).grid(row=3, column=5, padx=10, pady=5, sticky="w")

    # Iterations
    tk.Label(file_section, text="Iterations:", bg=THEME["card_bg"], fg=THEME["fg"],
            font=("Segoe UI", 10)).grid(row=4, column=0, padx=5, pady=10, sticky="w")
    iterations_entry = tk.Entry(file_section, width=10, bg=THEME["entry_bg"], fg=THEME["entry_fg"],
                               font=("Segoe UI", 10), relief="solid", bd=1)
    iterations_entry.grid(row=4, column=1, padx=5, pady=10, sticky="w")
    iterations_entry.insert(0, "10")

    # Log Display Card
    log_card = tk.Frame(main_container, bg=THEME["card_bg"], relief="solid", bd=1, padx=20, pady=20)
    log_card.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

    log_title = tk.Label(log_card, text="📝 Calibration Log",
                        font=("Segoe UI", 14, "bold"),
                        bg=THEME["card_bg"], fg=THEME["fg"])
    log_title.pack(anchor="w", pady=(0, 10))

    log_frame = tk.Frame(log_card, bg=THEME["card_bg"])
    log_frame.pack(fill=tk.BOTH, expand=True)

    log_display = tk.Text(log_frame, height=15, bg=THEME["log_bg"], fg=THEME["log_fg"],
                         wrap=tk.WORD, font=("Consolas", 9), relief="solid", bd=1)

    log_scrollbar = tk.Scrollbar(log_frame, orient="vertical", command=log_display.yview)
    log_display.configure(yscrollcommand=log_scrollbar.set)

    log_display.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Progress Section
    progress_frame = tk.Frame(main_container, bg=THEME["bg"])
    progress_frame.pack(fill=tk.X, pady=10)

    progress_label = tk.Label(progress_frame, text="0.0%", bg=THEME["bg"], fg=THEME["primary"],
                             font=("Segoe UI", 12, "bold"))
    progress_label.pack(side=tk.RIGHT, padx=10)

    progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", length=600, mode="determinate")
    progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

    # Action Buttons
    action_frame = tk.Frame(main_container, bg=THEME["bg"])
    action_frame.pack(fill=tk.X, pady=20)

    create_styled_button(action_frame, "🚀 Start Calibration", start_calibration,
                   style_type="success", width=20).pack(side=tk.LEFT, padx=5)

    create_styled_button(action_frame, "⏹️ Stop Calibration", stop_calibration,
                   style_type="danger", width=18).pack(side=tk.LEFT, padx=5)

    create_styled_button(action_frame, "📊 Show Results", open_results_directory,
                   style_type="info", width=15).pack(side=tk.LEFT, padx=5)



    create_styled_button(action_frame, "❌ Exit", exit_application,
                   style_type="danger", width=10).pack(side=tk.RIGHT, padx=5)

    # Apply initial zoom
    apply_zoom()

    # Show keyboard shortcuts info
    log_message("Keyboard shortcuts: Ctrl++ (Zoom In), Ctrl+- (Zoom Out), Ctrl+0 (Reset Zoom)")
    log_message("F1: Start Calibration, Esc: Exit Application")

    root.mainloop()

if __name__ == "__main__":
    if is_already_running():
        show_error("Another instance of the application is already running.")
        sys.exit(1)
    create_ui()
