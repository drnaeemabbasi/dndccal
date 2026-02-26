import os
import sys
import pandas as pd
import numpy as np
from skopt import gp_minimize
from skopt.space import Real
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import PatternFill
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.linear_model import LinearRegression
import subprocess
import shutil
import copy
import json
import logging
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import ctypes
import threading
import portalocker
from datetime import datetime

# --------------------- Path Handling for PyInstaller ---------------------
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Global variables
stop_calibration_flag = False
calibration_thread = None

# =====================================================================
#  SCALING ENGINE
#  All dimensions flow through S()/SF() so the entire UI scales uniformly.
#  base_dpi_scale: detected from OS (1.0 at 96dpi, 1.5 at 144dpi, 2.0 at 4K)
#  user_zoom: manual Ctrl+/- zoom (0.5x to 2.5x)
# =====================================================================
base_dpi_scale = 1.0
user_zoom = 1.0

def _detect_dpi_scale():
    """Detect OS DPI scaling BEFORE creating any tk window."""
    global base_dpi_scale
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)  # Per-monitor DPI v2
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass
    try:
        hdc = ctypes.windll.user32.GetDC(0)
        dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # LOGPIXELSX
        ctypes.windll.user32.ReleaseDC(0, hdc)
        base_dpi_scale = dpi / 96.0
    except Exception:
        base_dpi_scale = 1.0

_detect_dpi_scale()  # Call immediately at import time

def _scale_factor():
    return base_dpi_scale * user_zoom

def S(val):
    """Scale a pixel value by DPI * zoom. Returns int."""
    return int(val * _scale_factor())

def SF(base_size):
    """Scale a font size. Minimum 7."""
    return max(7, int(base_size * _scale_factor()))

def F(key):
    """Get a scaled font tuple by name."""
    fonts = {
        "heading_xl":  ("Segoe UI", SF(22), "bold"),
        "heading_lg":  ("Segoe UI", SF(16), "bold"),
        "heading_md":  ("Segoe UI", SF(13), "bold"),
        "heading_sm":  ("Segoe UI", SF(11), "bold"),
        "body":        ("Segoe UI", SF(10)),
        "body_sm":     ("Segoe UI", SF(9)),
        "mono":        ("Cascadia Code", SF(9)),
        "mono_sm":     ("Cascadia Code", SF(8)),
        "label":       ("Segoe UI", SF(10)),
        "label_bold":  ("Segoe UI", SF(10), "bold"),
        "button":      ("Segoe UI", SF(10), "bold"),
        "button_sm":   ("Segoe UI", SF(9), "bold"),
        "tag":         ("Segoe UI", SF(8), "bold"),
    }
    return fonts.get(key, fonts["body"])

# Widget registry for zoom rescaling
_scalable_widgets = []

def _register(widget):
    """Register a custom widget for rescaling on zoom."""
    _scalable_widgets.append(widget)
    return widget

def _labeled(parent, text, font_tag, **kw):
    """Create a Label with a font tag for zoom rescaling."""
    lbl = tk.Label(parent, text=text, font=F(font_tag), **kw)
    lbl._font_tag = font_tag
    return lbl

# =====================================================================
#  ZOOM FUNCTIONS (Ctrl+/- and Ctrl+0)
# =====================================================================
def zoom_in(event=None):
    global user_zoom
    user_zoom = min(2.5, round(user_zoom + 0.1, 1))
    _apply_zoom()

def zoom_out(event=None):
    global user_zoom
    user_zoom = max(0.5, round(user_zoom - 0.1, 1))
    _apply_zoom()

def reset_zoom(event=None):
    global user_zoom
    user_zoom = 1.0
    _apply_zoom()

def _apply_zoom():
    """Walk entire widget tree and update all fonts + rescale custom widgets."""
    log_message(f"  Zoom: {user_zoom:.1f}x  ·  DPI: {base_dpi_scale:.2f}x  ·  Effective: {_scale_factor():.2f}x")
    _walk_rescale(root)
    for w in _scalable_widgets:
        try:
            w.rescale()
        except Exception:
            pass

def _walk_rescale(widget):
    """Recursively update fonts on standard tk widgets."""
    try:
        if isinstance(widget, tk.Label) and hasattr(widget, '_font_tag'):
            widget.config(font=F(widget._font_tag))
        elif isinstance(widget, tk.Entry):
            widget.config(font=F("body"))
        elif isinstance(widget, tk.Text):
            widget.config(font=F("mono"))
        elif isinstance(widget, ttk.Combobox):
            widget.config(font=F("body"))
    except Exception:
        pass
    try:
        for child in widget.winfo_children():
            _walk_rescale(child)
    except Exception:
        pass


# =====================================================================
#  DESIGN SYSTEM  — Vercel/Linear inspired dark theme
# =====================================================================


# =====================================================================
#  DUAL THEME SYSTEM — Dark (default) + Light
# =====================================================================

DARK_THEME = {
    "bg_primary":     "#10131a",
    "bg_secondary":   "#181c27",
    "bg_tertiary":    "#212636",
    "bg_input":       "#1a1e2a",
    "bg_log":         "#0c0e14",
    "border":         "#2e3446",
    "border_focus":   "#3b82f6",
    "border_hover":   "#3d4456",
    "border_input":   "#363c50",
    "text_primary":   "#edf0f7",
    "text_secondary": "#a0a8be",
    "text_tertiary":  "#6b7280",
    "text_on_accent": "#ffffff",
    "accent":         "#06b6d4",
    "accent_hover":   "#0891b2",
    "accent_muted":   "#164e63",
    "success":        "#10b981",
    "success_hover":  "#059669",
    "danger":         "#ef4444",
    "danger_hover":   "#dc2626",
    "toggle_on":      "#10b981",
    "toggle_off":     "#3d4456",
    "toggle_knob":    "#ffffff",
    "progress_bg":    "#252a38",
    "progress_fill":  "#06b6d4",
}

LIGHT_THEME = {
    "bg_primary":     "#f0f2f5",
    "bg_secondary":   "#ffffff",
    "bg_tertiary":    "#e8ebf0",
    "bg_input":       "#f7f8fa",
    "bg_log":         "#fafbfc",
    "border":         "#d1d5db",
    "border_focus":   "#3b82f6",
    "border_hover":   "#b0b6c2",
    "border_input":   "#c8cdd6",
    "text_primary":   "#1a1d27",
    "text_secondary": "#4b5563",
    "text_tertiary":  "#9ca3af",
    "text_on_accent": "#ffffff",
    "accent":         "#0891b2",
    "accent_hover":   "#0e7490",
    "accent_muted":   "#cffafe",
    "success":        "#059669",
    "success_hover":  "#047857",
    "danger":         "#dc2626",
    "danger_hover":   "#b91c1c",
    "toggle_on":      "#10b981",
    "toggle_off":     "#d1d5db",
    "toggle_knob":    "#ffffff",
    "progress_bg":    "#e5e7eb",
    "progress_fill":  "#0891b2",
}

# Active palette — starts as dark
COLORS = dict(DARK_THEME)
_is_dark = True

def toggle_theme():
    """Switch between dark and light themes."""
    global _is_dark
    _is_dark = not _is_dark
    theme = DARK_THEME if _is_dark else LIGHT_THEME
    COLORS.update(theme)
    _apply_theme()

def _apply_theme():
    """Rebuild all widget colors after theme change."""
    _walk_retheme(root)
    # Rescale also redraws custom canvas widgets with new colors
    for w in _scalable_widgets:
        try:
            w.rescale()
        except Exception:
            pass
    # Re-style combobox
    sty = ttk.Style()
    sty.configure("Modern.TCombobox",
                  fieldbackground=COLORS["bg_input"], background=COLORS["bg_tertiary"],
                  foreground=COLORS["text_primary"], arrowcolor=COLORS["text_secondary"],
                  bordercolor=COLORS["border_input"])
    sty.map("Modern.TCombobox",
            fieldbackground=[("readonly", COLORS["bg_input"])],
            foreground=[("readonly", COLORS["text_primary"])])
    log_message(f"  Theme: {'Dark' if _is_dark else 'Light'}")

def _walk_retheme(widget):
    """Recursively recolor standard widgets."""
    try:
        wtype = widget.winfo_class()
        if wtype == 'Frame':
            # Determine if it's a card (bg_secondary) or main (bg_primary)
            try:
                old_bg = widget.cget("bg")
            except:
                old_bg = ""
            if old_bg in (DARK_THEME["bg_secondary"], LIGHT_THEME["bg_secondary"]):
                widget.config(bg=COLORS["bg_secondary"])
            elif old_bg in (DARK_THEME["accent_muted"], LIGHT_THEME["accent_muted"]):
                widget.config(bg=COLORS["accent_muted"])
            else:
                widget.config(bg=COLORS["bg_primary"])
        elif wtype == 'Label':
            parent_bg = COLORS["bg_primary"]
            try:
                parent_bg = widget.master.cget("bg")
            except:
                pass
            widget.config(bg=parent_bg)
            # Determine text color from font tag
            fg = widget.cget("fg")
            if fg in (DARK_THEME["text_primary"], LIGHT_THEME["text_primary"]):
                widget.config(fg=COLORS["text_primary"])
            elif fg in (DARK_THEME["text_secondary"], LIGHT_THEME["text_secondary"]):
                widget.config(fg=COLORS["text_secondary"])
            elif fg in (DARK_THEME["text_tertiary"], LIGHT_THEME["text_tertiary"]):
                widget.config(fg=COLORS["text_tertiary"])
            elif fg in (DARK_THEME["accent"], LIGHT_THEME["accent"]):
                widget.config(fg=COLORS["accent"])
        elif wtype == 'Entry':
            widget.config(bg=COLORS["bg_input"], fg=COLORS["text_primary"],
                         insertbackground=COLORS["accent"],
                         highlightbackground=COLORS["border_input"],
                         highlightcolor=COLORS["border_focus"])
        elif wtype == 'Text':
            widget.config(bg=COLORS["bg_log"], fg=COLORS["text_secondary"],
                         insertbackground=COLORS["accent"],
                         highlightbackground=COLORS["border_input"])
    except Exception:
        pass
    try:
        for child in widget.winfo_children():
            _walk_retheme(child)
    except Exception:
        pass


# =====================================================================
#  CUSTOM WIDGETS — Polished, DPI-aware, visible borders
# =====================================================================

class ModernCard(tk.Frame):
    def __init__(self, parent, title=None, icon=None, **kwargs):
        super().__init__(parent, bg=COLORS["bg_secondary"],
                         highlightbackground=COLORS["border"], highlightthickness=1, **kwargs)
        if title:
            header = tk.Frame(self, bg=COLORS["bg_secondary"])
            header.pack(fill=tk.X, padx=S(20), pady=(S(14), S(6)))
            text = f"{icon}  {title}" if icon else title
            _labeled(header, text, "heading_md",
                     bg=COLORS["bg_secondary"], fg=COLORS["text_primary"]).pack(side=tk.LEFT)


class ModernButton(tk.Canvas):
    """Rounded button with proper text sizing and visible borders."""
    def __init__(self, parent, text, command=None, style="primary",
                 width=140, height=38, icon=None, **kwargs):
        self._base_w = width
        self._base_h = height
        w, h = S(width), S(height)
        super().__init__(parent, width=w, height=h,
                         bg=parent.cget("bg"), highlightthickness=0, **kwargs)
        self._command = command
        self._icon = icon
        self._label = text
        self._text = f"{icon} {text}" if icon else text
        self._style_name = style
        self._width = w
        self._height = h
        self._hovered = False
        self._pressed = False

        styles = {
            "primary":   (COLORS["accent"],      COLORS["accent_hover"],  "#ffffff",             ""),
            "success":   (COLORS["success"],      COLORS["success_hover"], "#ffffff",             ""),
            "danger":    (COLORS["danger"],       COLORS["danger_hover"],  "#ffffff",             ""),
            "ghost":     (COLORS["bg_tertiary"],  COLORS["border_hover"],  COLORS["text_primary"], COLORS["border"]),
            "outline":   ("transparent",          COLORS["bg_tertiary"],   COLORS["accent"],      COLORS["accent"]),
            "secondary": (COLORS["bg_tertiary"],  COLORS["border_hover"],  COLORS["text_primary"], COLORS["border"]),
        }
        self._bg, self._hover_bg, self._fg, self._outline = styles.get(style, styles["primary"])

        self._draw()
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<ButtonPress-1>", self._on_press)
        self.bind("<ButtonRelease-1>", self._on_release)

    def _draw(self):
        self.delete("all")
        r = S(6)
        w, h = self._width, self._height
        bg = self._hover_bg if self._hovered else self._bg
        if bg == "transparent":
            bg = self.master.cget("bg")
        outline = self._outline if (self._outline and not self._hovered) else ""
        self._rrect(1, 1, w-1, h-1, r, fill=bg, outline=outline, width=1)
        # Font that won't overflow: use smaller font for compact buttons
        font = F("button") if self._base_w >= 120 else F("button_sm")
        self.create_text(w/2, h/2, text=self._text, font=font, fill=self._fg)

    def _rrect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
               x2, y2-r, x2, y2, x2-r, y2, x1+r, y2,
               x1, y2, x1, y2-r, x1, y1+r, x1, y1]
        return self.create_polygon(pts, smooth=True, **kw)

    def _on_enter(self, e):
        self._hovered = True; self._draw(); self.config(cursor="hand2")
    def _on_leave(self, e):
        self._hovered = False; self._draw()
    def _on_press(self, e):
        self._pressed = True
    def _on_release(self, e):
        if self._pressed and self._hovered and self._command: self._command()
        self._pressed = False

    def rescale(self):
        self._width, self._height = S(self._base_w), S(self._base_h)
        self.config(width=self._width, height=self._height)
        # Re-read colors for theme switch
        styles = {
            "primary":   (COLORS["accent"],      COLORS["accent_hover"],  "#ffffff",             ""),
            "success":   (COLORS["success"],      COLORS["success_hover"], "#ffffff",             ""),
            "danger":    (COLORS["danger"],       COLORS["danger_hover"],  "#ffffff",             ""),
            "ghost":     (COLORS["bg_tertiary"],  COLORS["border_hover"],  COLORS["text_primary"], COLORS["border"]),
            "outline":   ("transparent",          COLORS["bg_tertiary"],   COLORS["accent"],      COLORS["accent"]),
            "secondary": (COLORS["bg_tertiary"],  COLORS["border_hover"],  COLORS["text_primary"], COLORS["border"]),
        }
        self._bg, self._hover_bg, self._fg, self._outline = styles.get(self._style_name, styles["primary"])
        try:
            self.config(bg=self.master.cget("bg"))
        except: pass
        self._draw()


class ModernEntry(tk.Frame):
    """Entry with visible border and focus ring."""
    def __init__(self, parent, width=45, placeholder="", **kwargs):
        super().__init__(parent, bg=parent.cget("bg"))
        self.entry = tk.Entry(
            self, width=width,
            bg=COLORS["bg_input"], fg=COLORS["text_primary"],
            insertbackground=COLORS["accent"], font=F("body"),
            relief="flat", bd=0,
            highlightbackground=COLORS["border_input"],
            highlightcolor=COLORS["border_focus"],
            highlightthickness=1,
        )
        self.entry.pack(padx=0, pady=0, ipady=S(5), ipadx=S(6))
        self._placeholder = placeholder
        self._has_placeholder = False
        if placeholder:
            self._show_placeholder()
            self.entry.bind("<FocusIn>", self._on_focus_in)
            self.entry.bind("<FocusOut>", self._on_focus_out)

    def _show_placeholder(self):
        self.entry.insert(0, self._placeholder)
        self.entry.config(fg=COLORS["text_tertiary"]); self._has_placeholder = True
    def _on_focus_in(self, e):
        if self._has_placeholder:
            self.entry.delete(0, tk.END); self.entry.config(fg=COLORS["text_primary"]); self._has_placeholder = False
    def _on_focus_out(self, e):
        if not self.entry.get(): self._show_placeholder()
    def get(self):
        return "" if self._has_placeholder else self.entry.get()
    def delete(self, *a): self.entry.delete(*a)
    def insert(self, *a):
        if self._has_placeholder:
            self.entry.delete(0, tk.END); self.entry.config(fg=COLORS["text_primary"]); self._has_placeholder = False
        self.entry.insert(*a)
    def rescale(self):
        self.entry.config(font=F("body"))


class ModernCombobox(ttk.Combobox):
    def __init__(self, parent, **kwargs):
        sty = ttk.Style()
        sty.configure("Modern.TCombobox",
                      fieldbackground=COLORS["bg_input"], background=COLORS["bg_tertiary"],
                      foreground=COLORS["text_primary"], arrowcolor=COLORS["text_secondary"],
                      bordercolor=COLORS["border_input"], lightcolor=COLORS["border_input"],
                      darkcolor=COLORS["border_input"],
                      selectbackground=COLORS["accent_muted"], selectforeground=COLORS["text_primary"],
                      padding=(S(8), S(5)))
        sty.map("Modern.TCombobox",
                fieldbackground=[("readonly", COLORS["bg_input"])],
                foreground=[("readonly", COLORS["text_primary"])],
                bordercolor=[("focus", COLORS["border_focus"])])
        super().__init__(parent, style="Modern.TCombobox", **kwargs)


class ModernProgressBar(tk.Canvas):
    def __init__(self, parent, height=6, **kwargs):
        self._base_h = height; h = S(height)
        super().__init__(parent, height=h, bg=parent.cget("bg"), highlightthickness=0, **kwargs)
        self._height = h; self._value = 0
        self.bind("<Configure>", lambda e: self._draw())
    def _draw(self):
        self.delete("all"); w = self.winfo_width(); h = self._height
        if w < 10: return
        r = h // 2
        self._rrect(0, 0, w, h, r, fill=COLORS["progress_bg"])
        if self._value > 0:
            self._rrect(0, 0, max(h, (self._value/100)*w), h, r, fill=COLORS["progress_fill"])
    def _rrect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
               x2, y2-r, x2, y2, x2-r, y2, x1+r, y2,
               x1, y2, x1, y2-r, x1, y1+r, x1, y1]
        return self.create_polygon(pts, smooth=True, **kw)
    def set_value(self, v):
        self._value = max(0, min(100, v)); self._draw()
    def rescale(self):
        self._height = S(self._base_h); self.config(height=self._height); self._draw()


class ModernToggle(tk.Canvas):
    """Toggle with GREEN on / GRAY off — clearly visible states."""
    def __init__(self, parent, text="", variable=None, **kwargs):
        self._frame = tk.Frame(parent, bg=parent.cget("bg"))
        tw, th = S(44), S(24)
        super().__init__(self._frame, width=tw, height=th,
                         bg=parent.cget("bg"), highlightthickness=0)
        self._var = variable or tk.BooleanVar(value=False)
        self._tw, self._th = tw, th
        self.pack(side=tk.LEFT)
        self._label = None
        if text:
            self._label = _labeled(self._frame, text, "body",
                                   bg=parent.cget("bg"), fg=COLORS["text_secondary"])
            self._label.pack(side=tk.LEFT, padx=(S(8), 0))
        self._draw()
        self.bind("<Button-1>", self._toggle)

    @property
    def frame(self):
        return self._frame

    def _draw(self):
        self.delete("all")
        on = self._var.get()
        w, h = self._tw, self._th
        r = h // 2
        # GREEN when on, visible GRAY when off
        track = COLORS["toggle_on"] if on else COLORS["toggle_off"]
        self._pill(1, 1, w-1, h-1, r, fill=track, outline="")
        # White knob
        knob_d = int(h * 0.7)
        margin = (h - knob_d) // 2
        kx = w - margin - knob_d if on else margin
        self.create_oval(kx, margin, kx+knob_d, margin+knob_d,
                        fill=COLORS["toggle_knob"], outline="")

    def _pill(self, x1, y1, x2, y2, r, **kw):
        pts = [x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
               x2, y2-r, x2, y2, x2-r, y2, x1+r, y2,
               x1, y2, x1, y2-r, x1, y1+r, x1, y1]
        return self.create_polygon(pts, smooth=True, **kw)

    def _toggle(self, e=None):
        self._var.set(not self._var.get()); self._draw(); self.config(cursor="hand2")

    def get(self):
        return self._var.get()

    def rescale(self):
        self._tw, self._th = S(44), S(24)
        self.config(width=self._tw, height=self._th); self._draw()
        if self._label: self._label.config(font=F("body"))





# =====================================================================
#  SOIL DEPTH CONFIGS
# =====================================================================
SOIL_TEMP_DEPTHS = {
    "1cm": 5, "5cm": 6, "10cm": 7, "20cm": 8, "30cm": 9, "40cm": 10,
    "50cm": 11, "60cm": 12, "70cm": 13, "80cm": 14, "90cm": 15, "100cm": 16,
    "110cm": 17, "120cm": 18, "130cm": 19, "140cm": 20, "150cm": 21,
    "160cm": 22, "170cm": 23, "180cm": 24, "190cm": 25, "200cm": 26
}

SOIL_MOISTURE_DEPTHS = {
    "1cm": 27, "5cm": 28, "10cm": 29, "20cm": 30, "30cm": 31, "40cm": 32,
    "50cm": 33, "60cm": 34, "70cm": 35, "80cm": 36, "90cm": 37, "100cm": 38,
    "110cm": 39, "120cm": 40, "130cm": 41, "140cm": 42, "150cm": 43,
    "160cm": 44, "170cm": 45, "180cm": 46, "190cm": 47, "200cm": 48
}

ROOT_FOLDER = r"C:\DNDC"

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# =====================================================================
#  DYNAMIC PATH BUILDER
# =====================================================================
def get_output_paths(root_folder, site_name):
    output_dir = os.path.join(root_folder, "output_files")
    return {
        "output_dir": output_dir,
        "batch_record_root": os.path.join(output_dir, "Record", "Batch"),
        "results_dir": os.path.join(root_folder, "calibration_results", site_name),
    }

def detect_dndc_output_folder(batch_record_root):
    """After DNDC runs, find the most recently modified subfolder in Record/Batch/."""
    if not os.path.exists(batch_record_root):
        return None
    subfolders = [
        os.path.join(batch_record_root, d)
        for d in os.listdir(batch_record_root)
        if os.path.isdir(os.path.join(batch_record_root, d))
    ]
    if not subfolders:
        return None
    # Return the most recently modified
    return max(subfolders, key=os.path.getmtime)

def get_modeled_paths(dndc_record_dir):
    """Build CSV paths from the detected DNDC output folder."""
    return {
        "modeled_yield_csv": os.path.join(dndc_record_dir, "Multi_year_summary.csv"),
        "modeled_soil_climate_csv": os.path.join(dndc_record_dir, "Day_SoilClimate_1.csv"),
        "modeled_climate_csv": os.path.join(dndc_record_dir, "Day_Climate_1.csv"),
        "modeled_nee_csv": os.path.join(dndc_record_dir, "Day_SoilC_1.csv"),
        "modeled_n2o_csv": os.path.join(dndc_record_dir, "Day_SoilN_1.csv"),
    }

def auto_detect_site_name(batch_file_path):
    try:
        if not os.path.exists(batch_file_path):
            return ""
        with open(batch_file_path, 'r') as f:
            content = f.read()
        for line in content.strip().split('\n'):
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if line.endswith('.dnd') or '-stc' in line:
                name = os.path.splitext(os.path.basename(line))[0]
                if name:
                    return name
        import re
        match = re.search(r'(\S+\.dnd)', content)
        if match:
            return os.path.splitext(os.path.basename(match.group(1)))[0]
    except Exception:
        pass
    return ""


# =====================================================================
#  UTILITY FUNCTIONS
# =====================================================================
def show_error(message):
    def _show():
        messagebox.showerror("Error", message)
    if threading.current_thread() is threading.main_thread():
        _show()
    else:
        root.after(0, _show)

def log_message(message):
    def _log():
        try:
            log_display.insert(tk.END, message + "\n")
            log_display.see(tk.END)
        except Exception:
            pass
    if threading.current_thread() is threading.main_thread():
        _log()
    else:
        root.after(0, _log)

def check_file_exists(file_path):
    if not os.path.exists(file_path):
        log_message(f"✗ File not found: {file_path}")
        return False
    if not os.access(file_path, os.R_OK):
        log_message(f"✗ File not readable: {file_path}")
        return False
    return True

def read_dnd_file(dnd_path):
    try:
        with open(dnd_path, 'r') as file:
            return file.readlines()
    except Exception as e:
        log_message(f"✗ Failed to read .dnd file: {e}")
        return []

def write_dnd_file(dnd_path, lines):
    try:
        with open(dnd_path, 'w') as file:
            file.writelines(lines)
    except Exception as e:
        log_message(f"✗ Failed to write .dnd file: {e}")

def read_param_ranges(param_csv):
    try:
        df = pd.read_csv(param_csv)
        required_cols = ["parameter_name", "min", "max", "line_number"]
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(f"Missing columns: {missing}")
        for _, row in df.iterrows():
            if row['min'] >= row['max']:
                raise ValueError(f"'{row['parameter_name']}': min >= max")
            if row['line_number'] < 0:
                raise ValueError(f"'{row['parameter_name']}': line_number < 1")
        return df
    except Exception as e:
        log_message(f"✗ Parameter CSV error: {e}")
        return pd.DataFrame()

def update_parameters(lines, param_values, param_ranges_df):
    """Update .dnd lines with new parameter values.
    param_values: list of values (same order as param_ranges_df rows)."""
    updated_lines = copy.deepcopy(lines)
    for i, (_, row) in enumerate(param_ranges_df.iterrows()):
        if i >= len(param_values):
            break
        value = param_values[i]
        line_idx = int(row["line_number"])
        if 0 <= line_idx < len(updated_lines):
            parts = updated_lines[line_idx].strip().split()
            if len(parts) >= 2:
                parts[1] = f"{value:.6f}"
                updated_lines[line_idx] = ' '.join(parts) + '\n'
        else:
            log_message(f"⚠ Line {line_idx + 1} out of range for '{row['parameter_name']}'")
    return updated_lines
    return updated_lines

def run_dndc(output_dir, root_folder, batch_file):
    dndc_exe = os.path.join(root_folder, "DNDC95.exe")
    if not os.path.exists(dndc_exe):
        raise FileNotFoundError(f"DNDC executable not found: {dndc_exe}")
    try:
        result = subprocess.run([
            dndc_exe, "-root", root_folder, "-output", output_dir,
            "-s", batch_file, "-daily", "1"
        ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
           text=True, timeout=600)
        log_message("  ✓ DNDC run completed")
    except subprocess.TimeoutExpired:
        log_message("✗ DNDC timed out (600s)")
        raise
    except subprocess.CalledProcessError as e:
        log_message(f"✗ DNDC failed: {e.stderr}")
        raise


# =====================================================================
#  DATA READING FUNCTIONS
# =====================================================================
def read_yield_data(modeled_path, observed_path):
    try:
        if not check_file_exists(modeled_path) or not check_file_exists(observed_path):
            return pd.DataFrame(), pd.DataFrame()
        modeled_df = pd.read_csv(modeled_path, skiprows=5, usecols=[0, 2], names=['Year', 'Yield'], header=None)
        modeled_df.columns = ['Year', 'Yield_MOD']
        observed_df = pd.read_csv(observed_path, skiprows=2, usecols=[0, 1], names=['Year', 'Yield'], header=None)
        observed_df.columns = ['Year', 'Yield_OBS']
        return modeled_df, observed_df
    except Exception as e:
        log_message(f"✗ Yield read error: {e}")
        return pd.DataFrame(), pd.DataFrame()

def read_soil_temp_data(modeled_path, observed_path, depth):
    try:
        if isinstance(depth, (int, float)):
            depth = next((k for k, v in SOIL_TEMP_DEPTHS.items() if v == int(depth)), None)
            if depth is None:
                raise ValueError(f"Invalid depth: {depth}")
        if depth not in SOIL_TEMP_DEPTHS:
            raise ValueError(f"Invalid depth '{depth}'")
        col_idx = SOIL_TEMP_DEPTHS[depth]
        if not check_file_exists(modeled_path) or not check_file_exists(observed_path):
            return pd.DataFrame(), pd.DataFrame()
        modeled_df = pd.read_csv(modeled_path, skiprows=4, header=0)
        if col_idx >= len(modeled_df.columns):
            raise IndexError(f"Column {col_idx} out of range")
        modeled_df = modeled_df.iloc[:, [0, 1, col_idx]]
        modeled_df.columns = ['Year', 'Day', 'SoilTemp_MOD']
        observed_df = pd.read_csv(observed_path, skiprows=2, header=None,
                                  usecols=[0, 1, 2], names=['Year', 'Day', 'SoilTemp_OBS'])
        return modeled_df.dropna(), observed_df.dropna()
    except Exception as e:
        log_message(f"✗ SoilTemp read error at {depth}: {e}")
        return pd.DataFrame(), pd.DataFrame()

def read_soil_moisture_data(modeled_path, observed_path, depth):
    try:
        if isinstance(depth, (int, float)):
            depth = next(
                (k for k, v in SOIL_MOISTURE_DEPTHS.items() if v == int(depth) or int(depth) == int(k.replace("cm", ""))),
                None)
            if depth is None:
                return pd.DataFrame(), pd.DataFrame()
        if depth not in SOIL_MOISTURE_DEPTHS:
            return pd.DataFrame(), pd.DataFrame()
        col_idx = SOIL_MOISTURE_DEPTHS[depth]
        if not check_file_exists(modeled_path) or not check_file_exists(observed_path):
            return pd.DataFrame(), pd.DataFrame()
        modeled_df = pd.read_csv(modeled_path, skiprows=4, header=0)
        if col_idx >= len(modeled_df.columns):
            raise IndexError(f"Column {col_idx} out of range")
        modeled_df = modeled_df.iloc[:, [0, 1, col_idx]]
        modeled_df.columns = ['Year', 'Day', 'SoilMoisture_MOD']
        observed_df = pd.read_csv(observed_path, skiprows=2, header=None,
                                  usecols=[0, 1, 2], names=['Year', 'Day', 'SoilMoisture_OBS'])
        for df in [modeled_df, observed_df]:
            df['Year'] = pd.to_numeric(df['Year'], errors='coerce').fillna(0).astype(int)
            df['Day'] = pd.to_numeric(df['Day'], errors='coerce').fillna(0).astype(int)
            df.iloc[:, 2] = pd.to_numeric(df.iloc[:, 2], errors='coerce')
        return modeled_df.dropna(), observed_df.dropna()
    except Exception as e:
        log_message(f"✗ SoilMoisture read error at {depth}: {e}")
        return pd.DataFrame(), pd.DataFrame()

def read_et_data(modeled_path, observed_path):
    try:
        if not check_file_exists(modeled_path) or not check_file_exists(observed_path):
            return pd.DataFrame(), pd.DataFrame()
        modeled_df = pd.read_csv(modeled_path, skiprows=[0], header=1)
        modeled_df = modeled_df.iloc[:, [0, 1, 9]]
        modeled_df.columns = ['Year', 'Day', 'ET_MOD']
        modeled_df['Year'] = modeled_df['Year'].astype(int)
        modeled_df['Day'] = modeled_df['Day'].astype(int)
        modeled_df['ET_MOD'] = pd.to_numeric(modeled_df['ET_MOD'], errors='coerce')
        observed_df = pd.read_csv(observed_path, skiprows=2, header=None,
                                  usecols=[0, 1, 2], names=['Year', 'Day', 'ET_OBS'])
        observed_df['Year'] = observed_df['Year'].astype(int)
        observed_df['Day'] = observed_df['Day'].astype(int)
        observed_df['ET_OBS'] = pd.to_numeric(observed_df['ET_OBS'], errors='coerce')
        return modeled_df.dropna(), observed_df.dropna()
    except Exception as e:
        log_message(f"✗ ET read error: {e}")
        return pd.DataFrame(), pd.DataFrame()

def read_nee_data(modeled_path, observed_path):
    try:
        if not check_file_exists(modeled_path) or not check_file_exists(observed_path):
            return pd.DataFrame(), pd.DataFrame()
        modeled_df = pd.read_csv(modeled_path, skiprows=1)
        modeled_df = modeled_df.iloc[:, [0, 1, 42]]
        modeled_df.columns = ['Year', 'Day', 'NEE_MOD']
        modeled_df['Year'] = modeled_df['Year'].astype(int)
        modeled_df['Day'] = modeled_df['Day'].astype(int)
        modeled_df['NEE_MOD'] = pd.to_numeric(modeled_df['NEE_MOD'], errors='coerce')
        observed_df = pd.read_csv(observed_path, skiprows=2, usecols=[0, 1, 2],
                                  names=['Year', 'Day', 'NEE'], header=None)
        observed_df.columns = ['Year', 'Day', 'NEE_OBS']
        observed_df['Year'] = observed_df['Year'].astype(int)
        observed_df['Day'] = observed_df['Day'].astype(int)
        observed_df['NEE_OBS'] = pd.to_numeric(observed_df['NEE_OBS'], errors='coerce')
        return modeled_df, observed_df
    except Exception as e:
        log_message(f"✗ NEE read error: {e}")
        return pd.DataFrame(), pd.DataFrame()


def read_n2o_data(modeled_path, observed_path):
    try:
        if not check_file_exists(modeled_path) or not check_file_exists(observed_path):
            return pd.DataFrame(), pd.DataFrame()
        # Day_SoilN_1.csv: rows 1-2 junk, row 3 header, rows 4-5 units, data from row 6
        modeled_df = pd.read_csv(modeled_path, header=2, skiprows=[3, 4])
        modeled_df = modeled_df.iloc[:, [0, 1, 35]]  # Year, Day, column AJ (index 35)
        modeled_df.columns = ['Year', 'Day', 'N2O_MOD']
        modeled_df['Year'] = modeled_df['Year'].astype(int)
        modeled_df['Day'] = modeled_df['Day'].astype(int)
        modeled_df['N2O_MOD'] = pd.to_numeric(modeled_df['N2O_MOD'], errors='coerce')
        # Observed: same format as NEE (2 header rows, then Year,Day,Value)
        observed_df = pd.read_csv(observed_path, skiprows=2, usecols=[0, 1, 2],
                                  names=['Year', 'Day', 'N2O'], header=None)
        observed_df.columns = ['Year', 'Day', 'N2O_OBS']
        observed_df['Year'] = observed_df['Year'].astype(int)
        observed_df['Day'] = observed_df['Day'].astype(int)
        observed_df['N2O_OBS'] = pd.to_numeric(observed_df['N2O_OBS'], errors='coerce')
        return modeled_df, observed_df
    except Exception as e:
        log_message(f"✗ N2O read error: {e}")
        return pd.DataFrame(), pd.DataFrame()


# =====================================================================
#  METRICS
# =====================================================================
def calculate_metrics(y_true, y_pred):
    r2 = r2_score(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    mae = mean_absolute_error(y_true, y_pred)
    mbe = np.mean(y_pred - y_true)
    mean_obs = np.mean(y_true)
    nrmse = (rmse / mean_obs * 100) if mean_obs != 0 else np.nan
    try:
        X = np.array(y_pred).reshape(-1, 1)
        model = LinearRegression()
        model.fit(X, y_true)
        lr_r2 = model.score(X, y_true)
    except Exception:
        lr_r2 = np.nan
    return {'R2': r2, 'LR_R2': lr_r2, 'RMSE': rmse, 'nRMSE': nrmse, 'MAE': mae, 'MBE': mbe}

def match_and_evaluate(modeled_df, observed_df, target_var, yield_conversion_factor=0.4):
    if modeled_df.empty or observed_df.empty:
        return None, pd.DataFrame()
    for col in ['Year', 'Day']:
        if col in modeled_df.columns:
            modeled_df[col] = modeled_df[col].astype(int)
        if col in observed_df.columns:
            observed_df[col] = observed_df[col].astype(int)
    if target_var == "Yield":
        if 'Yield_MOD' in modeled_df.columns and yield_conversion_factor > 0:
            modeled_df['Yield_MOD'] = modeled_df['Yield_MOD'] / yield_conversion_factor
        merge_on = ['Year']
    else:
        merge_on = ['Year', 'Day']
    merged_df = pd.merge(modeled_df, observed_df, on=merge_on, how='inner')
    if merged_df.empty:
        return None, pd.DataFrame()
    observed_col = f"{target_var}_OBS"
    merged_df.dropna(subset=[observed_col], inplace=True)
    if merged_df.empty:
        return None, pd.DataFrame()
    metrics = calculate_metrics(merged_df[observed_col], merged_df[f"{target_var}_MOD"])
    return metrics, merged_df


# =====================================================================
#  CHECKPOINT SAVING
# =====================================================================
def save_iteration_outputs(results_dir, iteration, dndc_record_dir):
    """Copy the entire DNDC output folder for this iteration."""
    iter_dir = os.path.join(results_dir, "iteration_outputs", f"iter_{iteration:04d}")
    try:
        if os.path.exists(dndc_record_dir):
            shutil.copytree(dndc_record_dir, iter_dir, dirs_exist_ok=True)
    except Exception as e:
        log_message(f"⚠ Failed to save iteration {iteration} outputs: {e}")


# =====================================================================
#  OPTIMIZATION CORE
# =====================================================================
def objective_function(params, param_ranges_df, lines, target_var, depth, paths, batch_file, dnd_file, observed_csv):
    updated_lines = update_parameters(lines, params, param_ranges_df)
    write_dnd_file(dnd_file, updated_lines)
    root_folder = root_folder_entry.get().strip()
    run_dndc(paths["output_dir"], root_folder, batch_file)

    # Detect where DNDC actually wrote its output
    dndc_dir = detect_dndc_output_folder(paths["batch_record_root"])
    if not dndc_dir:
        log_message("✗ Could not find DNDC output folder")
        return np.inf
    mp = get_modeled_paths(dndc_dir)

    reader_map = {
        "Yield":        lambda: read_yield_data(mp["modeled_yield_csv"], observed_csv),
        "SoilTemp":     lambda: read_soil_temp_data(mp["modeled_soil_climate_csv"], observed_csv, depth),
        "SoilMoisture": lambda: read_soil_moisture_data(mp["modeled_soil_climate_csv"], observed_csv, depth),
        "ET":           lambda: read_et_data(mp["modeled_climate_csv"], observed_csv),
        "NEE":          lambda: read_nee_data(mp["modeled_nee_csv"], observed_csv),
        "N2O":          lambda: read_n2o_data(mp["modeled_n2o_csv"], observed_csv),
    }

    reader = reader_map.get(target_var)
    if not reader:
        return np.inf

    modeled_df, observed_df = reader()
    metrics, _ = match_and_evaluate(modeled_df, observed_df, target_var)
    if metrics is None:
        return np.inf

    return metrics['RMSE']


def bayesian_optimization(param_ranges, param_ranges_df, lines, target_var, depth,
                          paths, batch_file, dnd_file, observed_csv,
                          save_dnd_backups, save_iter_results):

    log_message(f"\n{'━'*50}")
    log_message(f"  Bayesian Optimization: {target_var}{f' @ {depth}' if depth else ''}")
    log_message(f"{'━'*50}")

    all_results = []
    best_rmse = np.inf
    best_params = None
    best_merged = None
    best_metrics = None
    best_iteration = 0
    iteration_counter = 0

    try:
        total_iterations = int(iterations_entry.get())
    except ValueError:
        total_iterations = 10

    results_dir = paths["results_dir"]
    os.makedirs(results_dir, exist_ok=True)

    dnd_backup_dir = os.path.join(results_dir, "dnd_backups")
    if save_dnd_backups:
        os.makedirs(dnd_backup_dir, exist_ok=True)

    def _read_after_run(observed_csv):
        """Detect DNDC output folder and build reader for current target."""
        dndc_dir = detect_dndc_output_folder(paths["batch_record_root"])
        if not dndc_dir:
            log_message("✗ Could not find DNDC output folder")
            return None, None, None
        mp = get_modeled_paths(dndc_dir)
        reader_map = {
            "Yield":        lambda: read_yield_data(mp["modeled_yield_csv"], observed_csv),
            "SoilTemp":     lambda: read_soil_temp_data(mp["modeled_soil_climate_csv"], observed_csv, depth),
            "SoilMoisture": lambda: read_soil_moisture_data(mp["modeled_soil_climate_csv"], observed_csv, depth),
            "ET":           lambda: read_et_data(mp["modeled_climate_csv"], observed_csv),
            "NEE":          lambda: read_nee_data(mp["modeled_nee_csv"], observed_csv),
            "N2O":          lambda: read_n2o_data(mp["modeled_n2o_csv"], observed_csv),
        }
        reader = reader_map.get(target_var)
        if not reader:
            return None, None, dndc_dir
        modeled_df, observed_df = reader()
        return modeled_df, observed_df, dndc_dir

    # ── Baseline run (iteration 0): original .dnd, no modifications ──
    log_message(f"\n  ⓪ Baseline: running with original parameters...")
    write_dnd_file(dnd_file, lines)
    root_folder = root_folder_entry.get().strip()
    run_dndc(paths["output_dir"], root_folder, batch_file)

    modeled_df, observed_df, dndc_dir = _read_after_run(observed_csv)
    if modeled_df is not None and not modeled_df.empty:
        baseline_metrics, baseline_merged = match_and_evaluate(modeled_df, observed_df, target_var)
        if baseline_metrics:
            # Extract actual parameter values from original .dnd
            original_params = []
            for _, row in param_ranges_df.iterrows():
                line_idx = int(row['line_number'])
                try:
                    parts = lines[line_idx].strip().split()
                    original_params.append(float(parts[1]) if len(parts) >= 2 else 0.0)
                except (IndexError, ValueError):
                    original_params.append(0.0)
            all_results.append({
                "Iteration": 0, "Parameters": original_params,
                "Metrics": baseline_metrics, "Merged_Data": baseline_merged
            })
            best_rmse = baseline_metrics['RMSE']
            best_params = original_params
            best_merged = baseline_merged
            best_metrics = baseline_metrics
            best_iteration = 0
            log_message(f"    R²={baseline_metrics['R2']:.4f}  RMSE={baseline_metrics['RMSE']:.2f}  "
                       f"nRMSE={baseline_metrics['nRMSE']:.1f}%")
            if save_dnd_backups:
                try: shutil.copy(dnd_file, os.path.join(dnd_backup_dir, "iter_0000_baseline.dnd"))
                except: pass
            if save_iter_results and dndc_dir:
                save_iteration_outputs(results_dir, 0, dndc_dir)
        else:
            log_message("    ⚠ Baseline produced no valid metrics")
    else:
        log_message("    ⚠ Baseline: no modeled data found")

    def callback(res):
        nonlocal best_rmse, best_params, best_merged, best_metrics, best_iteration, iteration_counter
        iteration_counter += 1

        global stop_calibration_flag
        if stop_calibration_flag:
            raise StopIteration("Stopped by user.")

        try:
            current_rmse = res.func_vals[-1]
            params = res.x_iters[-1]

            modeled_df, observed_df, dndc_dir = _read_after_run(observed_csv)
            if modeled_df is None or modeled_df.empty:
                return

            yield_metrics, merged_df = match_and_evaluate(modeled_df, observed_df, target_var)
            if yield_metrics is None:
                return

            all_results.append({
                "Iteration": iteration_counter, "Parameters": params,
                "Metrics": yield_metrics, "Merged_Data": merged_df
            })

            is_new_best = current_rmse < best_rmse
            if is_new_best:
                best_rmse = current_rmse
                best_params = params
                best_merged = merged_df
                best_metrics = yield_metrics
                best_iteration = iteration_counter

            marker = "★" if is_new_best else "·"
            log_message(f"\n  {marker} Iteration {iteration_counter}/{total_iterations}")
            log_message(f"    R²={yield_metrics['R2']:.4f}  RMSE={yield_metrics['RMSE']:.2f}  "
                       f"nRMSE={yield_metrics['nRMSE']:.1f}%  MAE={yield_metrics['MAE']:.2f}")

            pct = (iteration_counter / total_iterations) * 100
            root.after(0, lambda p=pct: progress_bar.set_value(p))
            root.after(0, lambda p=pct: progress_label.config(text=f"{p:.0f}%"))

            if save_dnd_backups:
                try: shutil.copy(dnd_file, os.path.join(dnd_backup_dir, f"iter_{iteration_counter:04d}.dnd"))
                except: pass

            if save_iter_results and dndc_dir:
                save_iteration_outputs(results_dir, iteration_counter, dndc_dir)

        except StopIteration:
            raise
        except Exception as e:
            log_message(f"  ✗ Iteration {iteration_counter} error: {e}")

    try:
        gp_minimize(
            lambda p: objective_function(p, param_ranges_df, lines, target_var, depth,
                                         paths, batch_file, dnd_file, observed_csv),
            param_ranges,
            n_calls=total_iterations,
            callback=callback,
            random_state=42,
            n_jobs=1
        )
        log_message("\n  ✓ Optimization complete.")
    except StopIteration:
        log_message("\n  ⏹ Stopped by user. Saving results...")

    return all_results, best_params, best_merged, best_metrics, best_iteration


# =====================================================================
#  RESULTS SAVING
# =====================================================================
def save_results(all_results, best_params, best_metrics, best_merged, best_iteration,
                 param_ranges_df, target_var, depth, results_dir):
    if not all_results:
        log_message("No results to save.")
        return

    os.makedirs(results_dir, exist_ok=True)

    if target_var in ["SoilTemp", "SoilMoisture"] and depth:
        output_file = os.path.join(results_dir, f"{target_var.lower()}_{depth}_results.xlsx")
    else:
        output_file = os.path.join(results_dir, f"{target_var.lower()}_calibration_results.xlsx")

    wb = Workbook()

    # All Iterations
    ws1 = wb.active
    ws1.title = "All Iterations"
    headers = ["Iteration"] + param_ranges_df['parameter_name'].tolist() + \
              ["R2", "LR_R2", "RMSE", "nRMSE(%)", "MAE", "MBE"]
    ws1.append(headers)

    best_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    for result in all_results:
        row = [result["Iteration"]] + list(result["Parameters"]) + [
            result["Metrics"]["R2"], result["Metrics"]["LR_R2"],
            result["Metrics"]["RMSE"], result["Metrics"]["nRMSE"],
            result["Metrics"]["MAE"], result["Metrics"]["MBE"]
        ]
        ws1.append(row)
        if result["Iteration"] == best_iteration:
            for cell in ws1[ws1.max_row]:
                cell.fill = best_fill

    # Best Iteration
    ws2 = wb.create_sheet("Best Iteration")
    ws2.append(["Best Iter #"] + param_ranges_df['parameter_name'].tolist() +
               ["R2", "LR_R2", "RMSE", "nRMSE(%)", "MAE", "MBE"])
    if best_params and best_metrics:
        ws2.append([best_iteration, *best_params,
                    best_metrics['R2'], best_metrics['LR_R2'],
                    best_metrics['RMSE'], best_metrics['nRMSE'],
                    best_metrics['MAE'], best_metrics['MBE']])

    # Data Comparison
    ws3 = wb.create_sheet("Data Comparison")
    if best_merged is not None and not best_merged.empty:
        has_day = 'Day' in best_merged.columns
        if has_day:
            ws3.append(["Year", "Day", "Observed", "Modeled"])
            for _, r in best_merged.iterrows():
                ws3.append([r['Year'], r['Day'], r[f"{target_var}_OBS"], r[f"{target_var}_MOD"]])
        else:
            ws3.append(["Year", "Observed", "Modeled"])
            for _, r in best_merged.iterrows():
                ws3.append([r['Year'], r[f"{target_var}_OBS"], r[f"{target_var}_MOD"]])

        ChartClass = LineChart if has_day else BarChart
        chart = ChartClass()
        chart.title = f"{target_var}{f' @ {depth}' if depth else ''} — Best Calibration"
        chart.style = 2
        chart.width = 28
        chart.height = 14
        data_col_start = 3 if has_day else 2
        data = Reference(ws3, min_col=data_col_start, max_col=data_col_start + 1,
                         min_row=1, max_row=len(best_merged) + 1)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=len(best_merged) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        if len(chart.series) >= 2:
            chart.series[0].graphicalProperties.solidFill = "4DC4E3"
            chart.series[1].graphicalProperties.solidFill = "FF6F61"
        # Legend outside plot area — bottom, no overlap
        from openpyxl.chart.layout import Layout, ManualLayout
        from openpyxl.chart.legend import Legend
        chart.legend = Legend()
        chart.legend.position = 'b'
        ws3.add_chart(chart, "F2")

    # Iteration Data — modeled values from each iteration side by side
    ws4 = wb.create_sheet("Iteration Data")
    if all_results:
        # Determine structure from first result's Merged_Data
        sample_merged = all_results[0].get("Merged_Data")
        if sample_merged is not None and not sample_merged.empty:
            has_day = 'Day' in sample_merged.columns
            mod_col = f"{target_var}_MOD"

            # Header row
            if has_day:
                header = ["Year", "Day"] + [f"Iteration {r['Iteration']}" for r in all_results]
            else:
                header = ["Year"] + [f"Iteration {r['Iteration']}" for r in all_results]
            ws4.append(header)

            # Build a dict of iteration data keyed by (Year,) or (Year, Day)
            # Use the merged data from each iteration
            all_keys = []
            iter_data = {}
            for result in all_results:
                md = result.get("Merged_Data")
                if md is None or md.empty:
                    continue
                it = result["Iteration"]
                iter_data[it] = {}
                for _, row in md.iterrows():
                    if has_day:
                        key = (int(row['Year']), int(row['Day']))
                    else:
                        key = (int(row['Year']),)
                    if key not in all_keys:
                        all_keys.append(key)
                    iter_data[it][key] = row[mod_col]

            # Sort keys
            all_keys.sort()

            # Write rows
            for key in all_keys:
                if has_day:
                    row = [key[0], key[1]]
                else:
                    row = [key[0]]
                for result in all_results:
                    it = result["Iteration"]
                    row.append(iter_data.get(it, {}).get(key, ""))
                ws4.append(row)

    # Save with graceful handling if file is locked/open
    for attempt in range(10):
        try:
            wb.save(output_file)
            log_message(f"  ✓ Results saved: {output_file}")
            break
        except PermissionError:
            if attempt == 0:
                log_message(f"  ⚠ Cannot write to {os.path.basename(output_file)} — file may be open in Excel.")
            # Try incremented filename
            base, ext = os.path.splitext(output_file)
            # Strip any existing _N suffix before adding new one
            import re
            base_clean = re.sub(r'_\d+$', '', base)
            output_file = f"{base_clean}_{attempt + 1}{ext}"
            if attempt == 9:
                log_message(f"  ✗ Could not save results after 10 attempts. Close Excel and re-run.")
            else:
                continue


# =====================================================================
#  CALIBRATION WORKFLOW
# =====================================================================
def start_calibration():
    global calibration_thread, stop_calibration_flag
    stop_calibration_flag = False

    if calibration_thread and calibration_thread.is_alive():
        log_message("⚠ A calibration is already running.")
        return

    rf = root_folder_entry.get().strip()
    bf = batch_file_entry.get().strip()
    df = dnd_file_entry.get().strip()
    oc = observed_csv_entry.get().strip()
    pc = param_csv_entry.get().strip()
    sn = site_name_entry.get().strip()

    if not all([rf, bf, df, oc, pc]):
        log_message("✗ Please fill all required fields.")
        return

    if not sn:
        sn = auto_detect_site_name(bf)
        site_name_entry.delete(0, tk.END)
        site_name_entry.insert(0, sn)
        log_message(f"  Auto-detected site: {sn}")

    try:
        n_iter = int(iterations_entry.get())
        if n_iter < 1:
            raise ValueError
    except ValueError:
        log_message("✗ Iterations must be a positive integer.")
        return

    for path, label in [(bf, "Batch"), (df, ".dnd"), (oc, "Observed CSV"), (pc, "Param CSV")]:
        if not os.path.exists(path):
            log_message(f"✗ {label} not found: {path}")
            return

    if not os.path.exists(os.path.join(rf, "DNDC95.exe")):
        log_message(f"✗ DNDC95.exe not found in: {rf}")
        return

    target_var = target_var_combo.get()
    depth = depth_combo.get() if target_var in ["SoilTemp", "SoilMoisture"] else None
    if target_var in ["SoilTemp", "SoilMoisture"] and not depth:
        log_message(f"✗ Select a depth for {target_var}.")
        return

    save_dnd = save_dnd_toggle.get()
    save_iter = save_checkpoint_toggle.get()

    log_message(f"\n{'═'*50}")
    log_message(f"  CALIBRATION START")
    log_message(f"  Target: {target_var}{f' @ {depth}' if depth else ''}")
    log_message(f"  Site: {sn}  |  Iterations: {n_iter}")
    log_message(f"  DND backups: {'on' if save_dnd else 'off'}  |  Save iteration results: {'on' if save_iter else 'off'}")
    log_message(f"{'═'*50}")

    progress_bar.set_value(0)
    progress_label.config(text="0%")

    calibration_thread = threading.Thread(
        target=calibrate_variable,
        args=(target_var, depth, rf, sn, bf, df, oc, pc, save_dnd, save_iter),
        daemon=True
    )
    calibration_thread.start()


def calibrate_variable(target_var, depth, root_folder, site_name,
                       batch_file, dnd_file, observed_csv, param_csv,
                       save_dnd_backups, save_iter_results):
    global stop_calibration_flag
    stop_calibration_flag = False

    paths = get_output_paths(root_folder, site_name)
    os.makedirs(paths["results_dir"], exist_ok=True)

    backup_path = dnd_file + ".backup"
    shutil.copy(dnd_file, backup_path)
    log_message(f"  ✓ .dnd backed up")

    try:
        lines = read_dnd_file(dnd_file)
        if not lines:
            log_message("✗ .dnd file empty or unreadable.")
            return

        param_ranges_df = read_param_ranges(param_csv)
        if param_ranges_df.empty:
            log_message("✗ Parameter CSV invalid.")
            return

        param_ranges = [(row['min'], row['max']) for _, row in param_ranges_df.iterrows()]

        results = bayesian_optimization(
            param_ranges, param_ranges_df, lines, target_var, depth,
            paths, batch_file, dnd_file, observed_csv,
            save_dnd_backups, save_iter_results
        )

        all_results, best_params, best_merged, best_metrics, best_iter = results

        if best_params and best_metrics:
            save_results(all_results, best_params, best_metrics, best_merged, best_iter,
                        param_ranges_df, target_var, depth, paths["results_dir"])
            log_message(f"\n  ✓ Best: Iteration #{best_iter}  RMSE={best_metrics['RMSE']:.4f}")
        else:
            log_message("⚠ No valid results found.")

    except Exception as e:
        log_message(f"✗ Calibration error: {e}")
        show_error(str(e))
    finally:
        shutil.copy(backup_path, dnd_file)
        log_message("  ✓ .dnd restored from backup")


def stop_calibration():
    global stop_calibration_flag
    if calibration_thread and calibration_thread.is_alive():
        def _confirm():
            global stop_calibration_flag
            if messagebox.askyesno("Stop", "Stop calibration? Results saved up to current iteration."):
                stop_calibration_flag = True
                log_message("  ⏹ Stopping after current iteration...")
        root.after(0, _confirm)
    else:
        log_message("No active calibration.")


# =====================================================================
#  UI EVENT HANDLERS
# =====================================================================
def on_target_var_change(event):
    tv = target_var_combo.get()
    if tv in ["SoilTemp", "SoilMoisture"]:
        depth_label.pack(side=tk.LEFT, padx=(0, S(8)), after=target_var_combo)
        depth_combo.pack(side=tk.LEFT, padx=(0, S(20)), after=depth_label)
        depth_combo['values'] = list(SOIL_TEMP_DEPTHS.keys() if tv == "SoilTemp" else SOIL_MOISTURE_DEPTHS.keys())
        depth_combo.current(0)
    else:
        depth_label.pack_forget()
        depth_combo.pack_forget()

def browse_file(entry_widget, file_type, title):
    f = filedialog.askopenfilename(filetypes=[(f"{file_type[1:].upper()} files", file_type)], title=title)
    if f:
        if isinstance(entry_widget, ModernEntry):
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, f)
        else:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, f)

def browse_directory(entry_widget, title):
    d = filedialog.askdirectory(title=title)
    if d:
        if isinstance(entry_widget, ModernEntry):
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, d)
        else:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, d)

def download_observed_template():
    try:
        save_dir = filedialog.askdirectory(title="Save Template To")
        if not save_dir:
            return
        tv = target_var_combo.get()
        if tv == "Yield":
            t = "Description: Observed yield\nUnits: kgC/ha/y\nYear,Value\n1,5000\n2,5200\n3,5100\n"
        else:
            t = "Description: Observed daily data\nUnits: see variable\nYear,Day,Value\n1,1,5.0\n1,2,5.2\n"
        fname = f"observed_{tv.lower()}_template.csv"
        with open(os.path.join(save_dir, fname), "w") as f:
            f.write(t)
        log_message(f"  ✓ Template saved: {fname}")
    except Exception as e:
        log_message(f"✗ Template error: {e}")

def download_param_template():
    try:
        save_dir = filedialog.askdirectory(title="Save Template To")
        if not save_dir:
            return
        with open(os.path.join(save_dir, "parameter_template.csv"), "w") as f:
            f.write("parameter_name,min,max,line_number\nparam1,0.1,1.0,10\nparam2,0.5,2.0,15\n")
        log_message("  ✓ Param template saved")
    except Exception as e:
        log_message(f"✗ Template error: {e}")

def open_results_directory():
    try:
        rf = root_folder_entry.get().strip()
        sn = site_name_entry.get().strip()
        if rf and sn:
            d = os.path.join(rf, "calibration_results", sn)
            if os.path.exists(d):
                os.startfile(d)
                return
        if rf:
            os.startfile(rf)
    except Exception as e:
        log_message(f"✗ Could not open results: {e}")

def exit_application():
    global calibration_thread, stop_calibration_flag
    if calibration_thread and calibration_thread.is_alive():
        if not messagebox.askyesno("Exit", "Calibration running. Exit anyway?"):
            return
        stop_calibration_flag = True
        calibration_thread.join(timeout=10)
    root.destroy()

def is_already_running():
    lock_path = os.path.join(ROOT_FOLDER, "dnccalibration.lock")
    lock_file = None
    try:
        lock_file = open(lock_path, 'w')
        portalocker.lock(lock_file, portalocker.LOCK_EX | portalocker.LOCK_NB)
        return False
    except (portalocker.exceptions.LockException, OSError):
        if lock_file:
            lock_file.close()
        return True


# ── Short aliases used by UI builder ──
_browse_file = browse_file
_browse_dir = browse_directory
_dl_obs_template = download_observed_template
_dl_param_template = download_param_template
_open_results = open_results_directory
_exit = exit_application


# =====================================================================
#  BUILD THE UI
# =====================================================================

def create_ui():
    global root, log_display, batch_file_entry, dnd_file_entry, observed_csv_entry
    global param_csv_entry, iterations_entry, progress_bar, progress_label
    global target_var_combo, depth_combo, depth_label
    global root_folder_entry, site_name_entry
    global save_dnd_toggle, save_checkpoint_toggle

    root = tk.Tk()
    root.title("DNDC Calibration Studio")
    root.geometry(f"{S(1100)}x{S(820)}")
    root.minsize(S(800), S(600))
    root.resizable(True, True)
    root.config(bg=COLORS["bg_primary"])

    # DPI awareness already set at module level
    # Keyboard shortcuts
    root.bind("<Control-plus>", lambda e: zoom_in())
    root.bind("<Control-equal>", lambda e: zoom_in())
    root.bind("<Control-minus>", lambda e: zoom_out())
    root.bind("<Control-0>", lambda e: reset_zoom())
    root.bind("<F1>", lambda e: start_calibration())
    root.bind("<Escape>", lambda e: exit_application())
    root.bind("<Control-t>", lambda e: toggle_theme())

    # TTK style
    sty = ttk.Style(); sty.theme_use('clam')
    sty.configure("Modern.TCombobox",
                  fieldbackground=COLORS["bg_input"], background=COLORS["bg_tertiary"],
                  foreground=COLORS["text_primary"], arrowcolor=COLORS["text_secondary"],
                  bordercolor=COLORS["border_input"], padding=(S(8), S(5)))
    sty.map("Modern.TCombobox",
            fieldbackground=[("readonly", COLORS["bg_input"])],
            foreground=[("readonly", COLORS["text_primary"])])

    # ── NO outer scroll — fixed layout, only log scrolls ──
    main = tk.Frame(root, bg=COLORS["bg_primary"])
    main.pack(fill=tk.BOTH, expand=True)

    pad = S(28)

    # ══════════ HEADER BAR ══════════
    hbar = tk.Frame(main, bg=COLORS["bg_primary"])
    hbar.pack(fill=tk.X, padx=pad, pady=(S(18), S(8)))

    # Left: title
    _labeled(hbar, "DNDC", "heading_xl",
             bg=COLORS["bg_primary"], fg=COLORS["accent"]).pack(side=tk.LEFT)
    _labeled(hbar, "  Calibration Studio", "heading_xl",
             bg=COLORS["bg_primary"], fg=COLORS["text_primary"]).pack(side=tk.LEFT)

    # Right side: theme toggle + zoom hint
    right_bar = tk.Frame(hbar, bg=COLORS["bg_primary"])
    right_bar.pack(side=tk.RIGHT)
    _labeled(right_bar, "Ctrl +/-  Zoom", "body_sm",
             bg=COLORS["bg_primary"], fg=COLORS["text_tertiary"]).pack(side=tk.RIGHT)
    # Theme toggle button
    theme_btn = ModernButton(right_bar, "◐ Theme", toggle_theme,
                             style="ghost", width=100, height=30)
    theme_btn.pack(side=tk.RIGHT, padx=(0, S(16))); _register(theme_btn)

    # Divider
    tk.Frame(main, bg=COLORS["border"], height=1).pack(fill=tk.X, padx=pad, pady=(S(4), S(10)))

    # ══════════ PROJECT CONFIG ══════════
    c1 = ModernCard(main, title="Project Configuration", icon="⚙")
    c1.pack(fill=tk.X, padx=pad, pady=(0, S(8)))
    g1 = tk.Frame(c1, bg=COLORS["bg_secondary"])
    g1.pack(fill=tk.X, padx=S(20), pady=(S(2), S(14)))

    # DNDC Root hardcoded — no UI row needed
    class _RootStub:
        def get(self): return ROOT_FOLDER
        def strip(self): return ROOT_FOLDER
    root_folder_entry = _RootStub()

    _labeled(g1, "Site Name", "label",
             bg=COLORS["bg_secondary"], fg=COLORS["text_secondary"]).pack(side=tk.LEFT, padx=(0, S(12)))
    site_name_entry = ModernEntry(g1, width=25, placeholder="e.g. F14-2024WEST")
    site_name_entry.pack(side=tk.LEFT)
    _labeled(g1, "auto-detected from batch file if blank", "body_sm",
             bg=COLORS["bg_secondary"], fg=COLORS["text_tertiary"]).pack(side=tk.LEFT, padx=(S(16), 0))

    # ══════════ CALIBRATION SETTINGS ══════════
    c2 = ModernCard(main, title="Calibration Settings", icon="◎")
    c2.pack(fill=tk.X, padx=pad, pady=(0, S(8)))
    g2 = tk.Frame(c2, bg=COLORS["bg_secondary"])
    g2.pack(fill=tk.X, padx=S(20), pady=(S(2), S(14)))

    # Row 0: Target + Depth only
    r0 = tk.Frame(g2, bg=COLORS["bg_secondary"])
    r0.pack(fill=tk.X, pady=(0, S(4)))

    _labeled(r0, "Target", "label", bg=COLORS["bg_secondary"], fg=COLORS["text_secondary"]).pack(side=tk.LEFT, padx=(0, S(8)))
    target_var_combo = ModernCombobox(r0, values=["Yield", "SoilTemp", "SoilMoisture", "ET", "NEE", "N2O"],
                                     width=14, state="readonly", font=F("body"))
    target_var_combo.pack(side=tk.LEFT, padx=(0, S(20)))
    target_var_combo.current(0)
    target_var_combo.bind("<<ComboboxSelected>>", on_target_var_change)

    depth_label = _labeled(r0, "Depth", "label", bg=COLORS["bg_secondary"], fg=COLORS["text_secondary"])
    depth_label.pack(side=tk.LEFT, padx=(0, S(8)))
    depth_combo = ModernCombobox(r0, width=8, state="readonly", font=F("body"))
    depth_combo.pack(side=tk.LEFT, padx=(0, S(20)))
    depth_label.pack_forget()
    depth_combo.pack_forget()

    # File rows: grid for alignment — NO weight=1 on entry col, use fixed width
    file_grid = tk.Frame(g2, bg=COLORS["bg_secondary"])
    file_grid.pack(fill=tk.X, pady=(0, S(4)))

    file_defs = [
        ("Batch File",    "*.txt", "Select Batch File",   None),
        (".dnd File",     "*.dnd", "Select .dnd File",    None),
        ("Observed CSV",  "*.csv", "Select Observed CSV",  "obs"),
        ("Parameter CSV", "*.csv", "Select Parameter CSV", "param"),
    ]

    entries = []
    for i, (label_text, ftype, title, tmpl) in enumerate(file_defs):
        _labeled(file_grid, label_text, "label",
                 bg=COLORS["bg_secondary"], fg=COLORS["text_secondary"]).grid(
            row=i, column=0, sticky="e", pady=S(3), padx=(0, S(10)))

        e = ModernEntry(file_grid, width=55)
        e.grid(row=i, column=1, pady=S(3), padx=(0, S(8)), sticky="w")
        entries.append(e)

        bb = ModernButton(file_grid, "Browse",
                          lambda ew=e, ft=ftype, t=title: _browse_file(ew, ft, t),
                          style="ghost", width=90, height=30)
        bb.grid(row=i, column=2, pady=S(3), padx=(0, S(4))); _register(bb)

        if tmpl == "obs":
            tb = ModernButton(file_grid, "↓ Template", _dl_obs_template,
                              style="outline", width=110, height=30)
            tb.grid(row=i, column=3, pady=S(3)); _register(tb)
        elif tmpl == "param":
            tb = ModernButton(file_grid, "↓ Template", _dl_param_template,
                              style="outline", width=110, height=30)
            tb.grid(row=i, column=3, pady=S(3)); _register(tb)

    batch_file_entry, dnd_file_entry, observed_csv_entry, param_csv_entry = entries

    # Options row: Iterations + toggles — BELOW file rows
    opts = tk.Frame(g2, bg=COLORS["bg_secondary"])
    opts.pack(fill=tk.X, pady=(S(6), 0))

    _labeled(opts, "Iterations", "label", bg=COLORS["bg_secondary"], fg=COLORS["text_secondary"]).pack(side=tk.LEFT, padx=(0, S(8)))
    iterations_entry = ModernEntry(opts, width=6)
    iterations_entry.pack(side=tk.LEFT)
    iterations_entry.insert(0, "10")

    _spacer = tk.Frame(opts, bg=COLORS["bg_secondary"], width=S(30))
    _spacer.pack(side=tk.LEFT)
    save_dnd_var = tk.BooleanVar(value=False)
    save_cp_var = tk.BooleanVar(value=True)
    t1 = ModernToggle(opts, text="Save .dnd backups", variable=save_dnd_var)
    t1.frame.pack(side=tk.LEFT, padx=(0, S(20))); _register(t1)
    save_dnd_toggle = t1
    t2 = ModernToggle(opts, text="Save iteration results", variable=save_cp_var)
    t2.frame.pack(side=tk.LEFT); _register(t2)
    save_checkpoint_toggle = t2

    # ══════════ OUTPUT LOG (this is the only scrollable part) ══════════
    c3 = ModernCard(main, title="Output Log", icon="▸")
    c3.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(0, S(8)))
    lc = tk.Frame(c3, bg=COLORS["bg_secondary"])
    lc.pack(fill=tk.BOTH, expand=True, padx=S(20), pady=(S(2), S(14)))

    log_display = tk.Text(lc, height=10,
                          bg=COLORS["bg_log"], fg=COLORS["text_secondary"],
                          insertbackground=COLORS["accent"],
                          font=F("mono"), wrap=tk.WORD,
                          relief="flat", bd=0,
                          highlightbackground=COLORS["border_input"],
                          highlightthickness=1,
                          padx=S(12), pady=S(8))
    ls = tk.Scrollbar(lc, orient="vertical", command=log_display.yview)
    log_display.configure(yscrollcommand=ls.set)
    log_display.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    ls.pack(side=tk.RIGHT, fill=tk.Y)

    # ══════════ BOTTOM BAR: Progress + Actions ══════════
    bottom = tk.Frame(main, bg=COLORS["bg_primary"])
    bottom.pack(fill=tk.X, padx=pad, pady=(S(4), S(16)))

    # Progress
    prow = tk.Frame(bottom, bg=COLORS["bg_primary"])
    prow.pack(fill=tk.X, pady=(0, S(10)))
    progress_label = _labeled(prow, "0%", "heading_sm",
                              bg=COLORS["bg_primary"], fg=COLORS["accent"])
    progress_label.pack(side=tk.RIGHT, padx=(S(10), 0))
    progress_bar = ModernProgressBar(prow, height=6)
    progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)
    _register(progress_bar)

    # Buttons
    brow = tk.Frame(bottom, bg=COLORS["bg_primary"])
    brow.pack(fill=tk.X)
    for text, cmd, sty, w, ico in [
        ("Start Calibration", start_calibration, "success", 220, "▶"),
        ("Stop",              stop_calibration,  "danger",  80,  "■"),
        ("Results",           _open_results,     "secondary", 110, "📂"),
    ]:
        b = ModernButton(brow, text, cmd, style=sty, width=w, height=40, icon=ico)
        b.pack(side=tk.LEFT, padx=(0, S(10))); _register(b)

    b_exit = ModernButton(brow, "Exit", _exit, style="ghost", width=80, height=40)
    b_exit.pack(side=tk.RIGHT); _register(b_exit)

    # ── Welcome ──
    log_message("  DNDC Calibration Studio")
    log_message(f"  DPI: {base_dpi_scale:.2f}x ({int(base_dpi_scale * 96)} DPI)")
    log_message(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    log_message("")
    log_message("  Ctrl++/Ctrl+-  Zoom  ·  Ctrl+0 Reset")
    log_message("  Ctrl+T  Toggle light/dark theme")
    log_message("  F1: Start  ·  Esc: Exit")
    log_message("  Output: <root>/calibration_results/<site>/")
    log_message("")

    root.mainloop()


# =====================================================================
#  ENTRY POINT
# =====================================================================
if __name__ == "__main__":
    if is_already_running():
        try:
            r = tk.Tk(); r.withdraw()
            messagebox.showerror("Error", "Another instance running."); r.destroy()
        except: print("Another instance running.")
        sys.exit(1)
    create_ui()